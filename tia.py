"""
tia.py — TIA Portal Openness Anbindung
STA Thread · Fehler · Logging · Session · HMI · Bibliothek · Executor
"""

# ═══════════════════════════════════════════════════════════════════════════════
# VERSION
# ═══════════════════════════════════════════════════════════════════════════════
VERSION      = "1.7.0"
VERSION_DATE = "2026-06-16"
VERSION_INFO = {
    "version":      VERSION,
    "date":         VERSION_DATE,
    "file":         __file__,
    "changes": [
        "1.7.0: get/set/export_plc_config — SPS-DeviceItem-Attribute lesen, schreiben, als Excel exportieren",
        "1.6.0: get/set/export_hmi_runtime_settings — Unified Runtime-Einstellungen lesen, schreiben, exportieren",
        "1.5.0: export_hw_config — Hardware-Konfiguration als Excel exportieren",
        "1.4.0: BUG-15 Fix: list_plc_tags comment-Feld via _mltext() korrekt auslesen",
        "1.3.0: list_plc_blocks — alle Bausteine inkl. Untergruppen, optionaler Gruppenfilter",
        "1.3.0: list_plc_tag_tables — alle Tag-Tabellen inkl. Untergruppen",
        "1.3.0: list_plc_tags — Tags einer Tabelle mit Typ, Adresse, Kommentar",
        "1.3.0: list_plc_udts — alle UDTs/Strukturen inkl. Untergruppen",
        "1.2.0: import_hmi_screen Screen-Name aus XML gelesen für korrektes Delete-vor-Import",
        "1.1.0: STA-Timeout 60s + Auto-Restart, export_hmi_tags Unified-Workaround, VBScriptFolder",
        "1.0.0: BUG-11–14 gefixt: rekursive Screen-Suche, Unified-Typ-Erkennung, HmiTarget-API",
    ]
}

import os, sys, threading, queue, logging, textwrap
from pathlib import Path
from typing import Any, Callable
from logging.handlers import RotatingFileHandler
from dataclasses import dataclass

# ═══════════════════════════════════════════════════════════════════════════════
# LOGGING
# ═══════════════════════════════════════════════════════════════════════════════

def _setup_logging(log_dir="C:/tia-mcp/logs", level="INFO"):
    Path(log_dir).mkdir(parents=True, exist_ok=True)
    log_file = Path(log_dir) / "tia_mcp.log"
    logging.basicConfig(
        level=getattr(logging, level.upper(), logging.INFO),
        format="%(asctime)s  %(levelname)-8s  %(name)-20s  %(message)s",
        datefmt="%Y-%m-%d %H:%M:%S",
        handlers=[
            RotatingFileHandler(log_file, maxBytes=5*1024*1024,
                                backupCount=3, encoding="utf-8"),
            logging.StreamHandler(sys.stderr),
        ]
    )

def _log(name): return logging.getLogger(f"tia.{name}")

# ═══════════════════════════════════════════════════════════════════════════════
# FEHLER
# ═══════════════════════════════════════════════════════════════════════════════

@dataclass
class TiaError(Exception):
    code: str
    message: str
    recoverable: bool
    details: dict | None = None

    def to_dict(self):
        return {"status":"error","code":self.code,"message":self.message,
                "recoverable":self.recoverable,
                **({"details":self.details} if self.details else {})}

_ERR_PATTERNS = [
    ("currently locked",  "PROJECT_LOCKED",   "Projekt gesperrt. Andere Session beenden.", False),
    ("No TIA Portal",     "NO_PORTAL_PROCESS","Kein TIA Portal Prozess. TIA Portal starten.", True),
    ("Access is denied",  "ACCESS_DENIED",    "Zugriff verweigert. Als Administrator starten.", False),
    ("already in use",    "PORTAL_IN_USE",    "TIA Portal durch andere Openness-Anwendung belegt.", False),
    ("FileNotFound",      "FILE_NOT_FOUND",   "Datei nicht gefunden. Pfad pruefen.", True),
    ("not found",         "OBJECT_NOT_FOUND", "Objekt nicht gefunden.", True),
]

def _translate(exc):
    msg = str(exc)
    for pattern, code, text, rec in _ERR_PATTERNS:
        if pattern.lower() in msg.lower():
            return TiaError(code, text, rec, {"original": msg})
    return TiaError("TIA_ERROR", f"TIA Fehler: {msg}", False,
                    {"type": type(exc).__name__, "original": msg})

def _tia_call(fn, *args, **kwargs):
    try:
        return fn(*args, **kwargs)
    except TiaError:
        raise
    except Exception as e:
        raise _translate(e) from e

# ═══════════════════════════════════════════════════════════════════════════════
# STA THREAD
# ═══════════════════════════════════════════════════════════════════════════════

_STA_TIMEOUT_DEFAULT  = 60   # Sekunden für normale Operationen
_STA_TIMEOUT_HEAVY    = 300  # Sekunden für schwere Ops: open_project, compile, close_portal

@dataclass
class _Job:
    fn: Callable; args: tuple; kwargs: dict; result_q: queue.Queue; timeout: float

class _Err:
    def __init__(self, e): self.exception = e

class STAThread:
    _instance = None; _lock = threading.Lock()

    def __new__(cls):
        with cls._lock:
            if not cls._instance:
                cls._instance = super().__new__(cls)
                cls._instance._started = False
                cls._instance._thread  = None
        return cls._instance

    def start(self):
        if self._started: return
        self._q: queue.Queue = queue.Queue()
        self._thread = threading.Thread(target=self._loop, name="TIA-STA", daemon=True)
        self._thread.start()
        self._started = True
        _log("sta").info("STA Thread gestartet")

    def stop(self):
        if not self._started: return
        self._q.put(None); self._started = False
        _log("sta").info("STA Thread gestoppt")

    def _restart(self):
        """STA-Thread neu starten nach Timeout. Session-Handles zurücksetzen."""
        _log("sta").warning("STA Thread neu gestartet nach Timeout.")
        self._started = False
        _sess.portal  = None
        _sess.project = None
        self._q = queue.Queue()
        self._thread = threading.Thread(target=self._loop, name="TIA-STA", daemon=True)
        self._thread.start()
        self._started = True

    def run(self, fn, *args, timeout=None, **kwargs):
        if not self._started: raise RuntimeError("STAThread nicht gestartet")
        t = timeout if timeout is not None else _STA_TIMEOUT_DEFAULT
        rq = queue.Queue()
        self._q.put(_Job(fn, args, kwargs, rq, t))
        try:
            out = rq.get(timeout=t)
        except queue.Empty:
            _log("sta").warning(f"STA Timeout nach {t}s — Thread wird neu gestartet.")
            self._restart()
            raise TiaError(
                "STA_TIMEOUT",
                f"TIA-Operation hat nach {t}s nicht geantwortet. "
                "STA-Thread wurde neu gestartet. "
                "Bitte connect_portal + attach_project erneut aufrufen.",
                True,
                {"hint": "STA thread restarted, session handles reset"}
            )
        if isinstance(out, _Err): raise out.exception
        return out

    def _loop(self):
        try:
            import pythoncom; pythoncom.CoInitialize()
        except ImportError: pass
        while True:
            job = self._q.get()
            if job is None: break
            try:
                job.result_q.put(job.fn(*job.args, **job.kwargs))
            except Exception as e:
                _log("sta").error(str(e), exc_info=True)
                job.result_q.put(_Err(e))
        try:
            import pythoncom; pythoncom.CoUninitialize()
        except ImportError: pass

sta = STAThread()

# ═══════════════════════════════════════════════════════════════════════════════
# SESSION
# ═══════════════════════════════════════════════════════════════════════════════

_TIA_BASE = r"C:\Program Files\Siemens\Automation"
_VERSIONS = ["V21","V20","V19","V18"]

class _Session:
    portal = None; project = None
    tia_version = None; dll_path = None; _dlls_loaded = False

    # DLLs die geladen werden muessen
    _V21_DLLS = [
        "Siemens.Engineering.Base.dll",      # Hauptassembly (TiaPortal, Project, ...)
        "Siemens.Engineering.Step7.dll",     # PLC (PlcSoftware, Blocks, Tags)
        "Siemens.Engineering.WinCC.dll",     # HMI Advanced
        "Siemens.Engineering.WinCCUnified.dll",  # HMI Unified
        "Siemens.Engineering.AddIn.Base.dll",    # AddIn-Basis
    ]
    _LEGACY_DLLS = [
        "Siemens.Engineering.dll",           # V19/V20 Monolith
        "Siemens.Engineering.Hmi.dll",
        "Siemens.Engineering.HmiUnified.dll",
    ]

    def ensure_dlls(self):
        if self._dlls_loaded: return
        # Pfad suchen — V21 erkennt man an Base.dll statt Engineering.dll
        for v in _VERSIONS:
            candidates = [
                Path(_TIA_BASE) / f"Portal {v}" / "PublicAPI" / v / "net48",  # V21
                Path(_TIA_BASE) / f"Portal {v}" / "PublicAPI" / v,            # V19/V20
                Path(_TIA_BASE) / f"Portal {v}" / "PublicAPI",                # Fallback
            ]
            for p in candidates:
                # V21: Base.dll, aeltere: Engineering.dll
                if (p / "Siemens.Engineering.Base.dll").exists() or                    (p / "Siemens.Engineering.dll").exists():
                    self.tia_version, self.dll_path = v, str(p); break
            if self.dll_path: break
        if not self.dll_path:
            raise TiaError("TIA_NOT_INSTALLED","Keine TIA Installation.",False,{"searched":_VERSIONS})

        import clr
        if self.dll_path not in sys.path: sys.path.append(self.dll_path)

        # Entscheiden ob V21 (aufgeteilte DLLs) oder aelter (Monolith)
        is_v21 = (Path(self.dll_path) / "Siemens.Engineering.Base.dll").exists()
        dlls = self._V21_DLLS if is_v21 else self._LEGACY_DLLS

        for dll in dlls:
            f = os.path.join(self.dll_path, dll)
            if os.path.exists(f):
                try:
                    clr.AddReference(f)
                    _log("session").debug(f"DLL geladen: {dll}")
                except Exception as e:
                    _log("session").warning(f"DLL nicht geladen: {dll} — {e}")

        _log("session").info(f"TIA {self.tia_version} DLLs geladen aus {self.dll_path}")
        self._dlls_loaded = True

    def ensure_portal(self):
        if not self.portal:
            raise TiaError("NOT_CONNECTED","Nicht verbunden. connect_portal aufrufen.",True)

    def ensure_project(self):
        self.ensure_portal()
        if not self.project:
            raise TiaError("NO_PROJECT","Kein Projekt. open_project aufrufen.",True)

_sess = _Session()

def connect_portal(mode="attach"):
    def _run():
        _sess.ensure_dlls()
        from Siemens.Engineering import TiaPortal, TiaPortalMode
        if mode == "attach":
            procs = list(TiaPortal.GetProcesses())
            if not procs: raise TiaError("NO_PORTAL_PROCESS","TIA Portal starten.",True)
            _sess.portal = procs[0].Attach()
            _log("session").info(f"Attached PID={procs[0].Id} V={_sess.tia_version}")
            return {"status":"ok","mode":"attach","tia_version":_sess.tia_version,"process_id":procs[0].Id}
        tm = TiaPortalMode.WithoutUserInterface if mode=="headless" else TiaPortalMode.WithUserInterface
        _sess.portal = TiaPortal(tm)
        _log("session").info(f"Gestartet mode={mode} V={_sess.tia_version}")
        return {"status":"ok","mode":mode,"tia_version":_sess.tia_version}
    return sta.run(_tia_call, _run)

def attach_project():
    """Holt das bereits in TIA Portal geoeffnete Projekt — kein Pfad noetig."""
    def _run():
        _sess.ensure_portal()
        projects = list(_sess.portal.Projects)
        if not projects:
            raise TiaError("NO_PROJECT",
                "Kein Projekt in TIA Portal offen. Bitte zuerst ein Projekt oeffnen.",True)
        _sess.project = projects[0]
        _log("session").info(f"Projekt uebernommen: {_sess.project.Name}")
        return {"status":"ok","project":_sess.project.Name,
                "path":str(_sess.project.Path)}
    return sta.run(_tia_call, _run)

def open_project(path, retries=10, retry_delay=10):
    """
    Projekt oeffnen (.ap21).
    retries/retry_delay: Wartet bis TIA Portal bereit ist — sinnvoll nach connect_portal(mode='gui').
    Standard: 10 Versuche x 10 Sekunden = max. 100 Sekunden.
    """
    import time
    def _run():
        _sess.ensure_portal()
        from System.IO import FileInfo
        fi = FileInfo(path)
        if not fi.Exists: raise TiaError("FILE_NOT_FOUND",f"Nicht gefunden: {path}",True)
        last_exc = None
        for attempt in range(max(1, retries)):
            try:
                _sess.project = _sess.portal.Projects.Open(fi)
                _log("session").info(f"Projekt: {_sess.project.Name} (Versuch {attempt+1})")
                return {"status":"ok","project":_sess.project.Name,"path":path}
            except Exception as e:
                last_exc = e
                _log("session").warning(f"open_project Versuch {attempt+1}/{retries} fehlgeschlagen: {e}")
                if attempt < retries - 1:
                    time.sleep(retry_delay)
        raise _translate(last_exc)
    return sta.run(_tia_call, _run, timeout=_STA_TIMEOUT_HEAVY)

def save_project():
    """Aktuelles Projekt speichern."""
    def _run():
        _sess.ensure_project()
        _sess.project.Save()
        _log("session").info(f"Projekt gespeichert: {_sess.project.Name}")
        return {"status":"ok","project":_sess.project.Name}
    return sta.run(_tia_call, _run, timeout=_STA_TIMEOUT_HEAVY)

def close_project():
    """
    Aktuelles Projekt schliessen (ohne Speichern).
    Vorher save_project() aufrufen um Aenderungen zu sichern.
    """
    def _run():
        _sess.ensure_project()
        name = _sess.project.Name
        _sess.project.Close()
        _sess.project = None
        _log("session").info(f"Projekt geschlossen: {name}")
        return {"status":"ok","closed":name}
    return sta.run(_tia_call, _run)

def close_portal():
    """
    TIA Portal beenden (Dispose + taskkill).
    Schliesst alle offenen Projekte und beendet den TIA-Prozess.
    Vorher save_project() und close_project() aufrufen.
    """
    import threading as _threading
    import subprocess as _subprocess

    if not _sess.portal:
        raise TiaError("NOT_CONNECTED", "Nicht verbunden. connect_portal aufrufen.", True)

    # PID vor dem Dispose merken — danach ist der Portal-Handle ggf. ungültig
    pid_holder = [None]
    def _get_pid():
        try:
            from Siemens.Engineering import TiaPortal
            procs = list(TiaPortal.GetProcesses())
            if procs:
                pid_holder[0] = procs[0].Id
        except Exception:
            pass
    try:
        sta.run(_get_pid)
    except Exception:
        pass

    # Dispose im STA-Thread mit 30s-Timeout
    done = _threading.Event()
    dispose_error = [None]

    def _dispose():
        try:
            if _sess.portal:
                _sess.portal.Dispose()
        except Exception as e:
            dispose_error[0] = e
        finally:
            _sess.portal  = None
            _sess.project = None
            done.set()

    rq = queue.Queue()
    sta._q.put(_Job(_dispose, (), {}, rq, 30))

    timed_out = not done.wait(timeout=30)
    if timed_out:
        _log("session").warning("close_portal: Dispose-Timeout nach 30s")
        _sess.portal  = None
        _sess.project = None

    if dispose_error[0]:
        _log("session").warning(f"close_portal Dispose-Fehler (ignoriert): {dispose_error[0]}")

    # Prozess killen — Dispose trennt nur die API-Verbindung, beendet TIA nicht immer
    killed = False
    if pid_holder[0]:
        try:
            _subprocess.run(
                ["taskkill", "/F", "/PID", str(pid_holder[0])],
                stdout=_subprocess.DEVNULL, stderr=_subprocess.DEVNULL,
                timeout=10
            )
            killed = True
            _log("session").info(f"TIA Portal Prozess {pid_holder[0]} beendet.")
        except Exception as ke:
            _log("session").error(f"taskkill fehlgeschlagen: {ke}")

    msg = "TIA Portal beendet."
    if timed_out:
        msg = "TIA Portal beendet (Dispose-Timeout, force kill)."
    elif killed:
        msg = "TIA Portal beendet."

    return {"status": "ok", "message": msg}

def get_session_status():
    def _run():
        ver = None
        for v in _VERSIONS:
            for sub in [f"PublicAPI/{v}/net48", f"PublicAPI/{v}"]:
                p = Path(_TIA_BASE) / f"Portal {v}" / sub
                if (p / "Siemens.Engineering.Base.dll").exists() or \
                   (p / "Siemens.Engineering.dll").exists():
                    ver = v; break
            if ver: break
        # project.Name nur aufrufen wenn Projekt bekannt — im STA-Thread ist das sicher
        proj_name = None
        if _sess.project:
            try:
                proj_name = _sess.project.Name
            except Exception:
                proj_name = None
        return {
            "tia_installed":    ver,
            "portal_connected": _sess.portal  is not None,
            "project_open":     _sess.project is not None,
            "project_name":     proj_name,
        }
    return sta.run(_tia_call, _run)

def get_project_info():
    def _run():
        _sess.ensure_project(); p = _sess.project
        return {"name":p.Name,"path":str(p.Path),
                "devices":[{"name":d.Name,"type":str(d.TypeIdentifier)}
                           for d in _iter_all_devices(p)]}
    return sta.run(_tia_call, _run)

def list_devices():
    def _run():
        _sess.ensure_project()
        import Siemens.Engineering as eng
        sw_type = _get_sw_container_type()
        result = []
        for device in _iter_all_devices(_sess.project):
            sw_list = []
            stack = list(device.DeviceItems)
            while stack:
                item = stack.pop()
                sw = _try_get_software(item, "", sw_type, eng)
                if sw:
                    full = type(sw).__module__ + "." + type(sw).__name__
                    hw = "Unified"  if "HmiUnified" in full else \
                         "Advanced" if "Hmi"        in full else \
                         "PLC"      if "Plc"        in full else type(sw).__name__
                    sw_list.append({"item": item.Name, "type": hw})
                try:
                    for sub in item.DeviceItems: stack.append(sub)
                except Exception: pass
            result.append({"name": device.Name, "software": sw_list})
        return {"devices": result, "count": len(result)}
    return sta.run(_tia_call, _run)

def _find_plc_item(device_name):
    """CPU DeviceItem anhand des Software-ItemNamens finden."""
    import Siemens.Engineering as eng
    sw_type = _get_sw_container_type()
    for device in _iter_all_devices(_sess.project):
        item_stack = list(device.DeviceItems)
        while item_stack:
            item = item_stack.pop()
            sw = _try_get_software(item, "", sw_type, eng)
            if sw and "Plc" in type(sw).__name__ and item.Name == device_name:
                return item
            for sub in item.DeviceItems:
                item_stack.append(sub)
    raise TiaError("PLC_NOT_FOUND", f"PLC '{device_name}' nicht gefunden.", True)


# Attribute die nicht sinnvoll schreibbar sind (interne/berechnete Werte)
_PLC_READONLY_ATTRS = {
    "Classification", "Container", "FirmwareVersion", "InstallationDate",
    "IsBuiltIn", "IsPlugged", "Items", "Name", "OrderNumber",
    "PositionNumber", "ShortDesignation", "TypeIdentifier",
    "TypeIdentifierNormalized", "TypeName", "MultilingualSupportAdvanced",
    "CommentML",
}

# Gruppierung für Excel-Export
_PLC_ATTR_GROUPS = {
    "Allgemein":     ["Name", "OrderNumber", "ShortDesignation", "FirmwareVersion",
                      "TypeName", "Author", "Comment", "LocationIdentifier",
                      "PlantDesignation", "AdditionalInformation", "InstallationDate"],
    "Zyklus":        ["CycleMinimumCycleTime", "CycleMaximumCycleTime",
                      "CycleCommunicationLoad", "CycleEnableMinimumCycleTime",
                      "IsochronousMode", "SendClock"],
    "Startup":       ["StartupActionAfterPowerOn", "StartupComparisonPresetToActualModule",
                      "StartupConfigurationTimeout"],
    "Zeitzone":      ["TimeOfDayLocalTimeZone", "TimeOfDayActivateDaylightSavingTime",
                      "TimeOfDayDaylightSavingTimeOffset", "TimeOfDayDaylightSavingTimeStartMonth",
                      "TimeOfDayDaylightSavingTimeStartWeek", "TimeOfDayDaylightSavingTimeStartWeekday",
                      "TimeOfDayDaylightSavingTimeStartHour", "TimeOfDayStandardTimeStartMonth",
                      "TimeOfDayStandardTimeStartWeek", "TimeOfDayStandardTimeStartWeekday",
                      "TimeOfDayStandardTimeStartHour", "TimeSynchronizationNtpV2"],
    "Sicherheit":    ["PlcAccessControlConfiguration", "ConfigurationControl",
                      "NetworkFaultsAsMaintenance", "SuppressDeactivatingSystemDiagnosticsAlarms",
                      "UseFixedSystemDiagnosticsAlarmIds", "ProtectionIntervalForSummarizeOfSecurityEvents",
                      "ProtectionSummarizeSecurityEventsOnHighLoad",
                      "ProtectionUnitForSummarizeOfSecurityEvents"],
    "Netzwerk":      ["HostAndDomainnameActive", "IPv4ForwardingActive", "CommunicationMode",
                      "PnDnsConfiguration", "PnDnsConfigNameResolve",
                      "SNMPActive", "SNMPReadOnlyActive", "SNMPReadOnlyCommunityName",
                      "SNMPReadWriteCommunityName", "SNMPSynchronizeActive", "SNMPConfigurationSource"],
    "OPC UA":        ["OpcUaPurchasedLicense"],
    "Web & Syslog":  ["WebserverActivate", "SysLogAutoAcceptClient",
                      "SysLogClientCertificateId", "SysLogTrustedCertificateIds"],
    "Speicher":      ["ClockMemoryByte", "SystemMemoryByte", "SystemPowerSupplyExternal"],
    "Diagnose":      ["CentralAlarmManagement", "DetectLoadVoltageFailure",
                      "ProDiagUsedLicenses"],
}


def get_plc_config(device_name):
    def _run():
        _sess.ensure_project()
        item = _find_plc_item(device_name)
        config = {}
        for ai in item.GetAttributeInfos():
            n = str(ai.Name)
            v = item.GetAttribute(n)
            config[n] = str(v) if v is not None else None
        return {"status": "ok", "device": device_name, "config": config}
    return sta.run(_tia_call, _run)


def set_plc_config(device_name, settings: dict):
    def _run():
        _sess.ensure_project()
        item = _find_plc_item(device_name)
        applied, skipped = [], []
        for key, val in settings.items():
            if key in _PLC_READONLY_ATTRS:
                skipped.append(key)
                continue
            item.SetAttribute(key, val)
            applied.append(key)
        return {"status": "ok", "device": device_name, "applied": applied, "skipped_readonly": skipped}
    return sta.run(_tia_call, _run)


def export_plc_config(device_name, output_path=None):
    def _run():
        _sess.ensure_project()
        item = _find_plc_item(device_name)
        config = {}
        for ai in item.GetAttributeInfos():
            n = str(ai.Name)
            v = item.GetAttribute(n)
            config[n] = str(v) if v is not None else None
        return config

    config = sta.run(_tia_call, _run)

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise TiaError("MISSING_DEPENDENCY", "openpyxl nicht installiert.", False)

    out = Path(output_path) if output_path else Path(_DEFAULT_EXPORT) / f"plc_config_{device_name}.xlsx"
    out.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "PLC Konfiguration"
    thin   = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col, (h, w) in enumerate(zip(["Einstellung", "Wert", "Gruppe"], [40, 40, 22]), 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font      = Font(name="Arial", bold=True, color="FFFFFF", size=10)
        c.fill      = PatternFill("solid", start_color="1F4E79")
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border    = border
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[1].height = 22

    grp_fill = PatternFill("solid", start_color="D6E4F0")
    val_fill = PatternFill("solid", start_color="F5FBFF")
    alt_fill = PatternFill("solid", start_color="EAF4FB")
    ro_font  = Font(name="Arial", size=10, color="808080")

    r = 2
    written = set()
    for group, keys in _PLC_ATTR_GROUPS.items():
        c = ws.cell(row=r, column=1, value=group)
        c.font = Font(name="Arial", bold=True, size=10)
        c.fill = grp_fill
        c.border = border
        for col in [2, 3]:
            ws.cell(row=r, column=col).fill   = grp_fill
            ws.cell(row=r, column=col).border = border
        r += 1
        for key in keys:
            if key not in config:
                continue
            fill = val_fill if r % 2 == 0 else alt_fill
            ro   = key in _PLC_READONLY_ATTRS
            ws.cell(row=r, column=1, value=f"  {key}").fill   = fill
            ws.cell(row=r, column=1).font   = ro_font if ro else Font(name="Arial", size=10)
            ws.cell(row=r, column=1).border = border
            ws.cell(row=r, column=2, value=config[key] or "").fill   = fill
            ws.cell(row=r, column=2).font   = Font(name="Arial", size=10)
            ws.cell(row=r, column=2).border = border
            ws.cell(row=r, column=3, value=group).fill   = fill
            ws.cell(row=r, column=3).font   = Font(name="Arial", size=10, color="808080")
            ws.cell(row=r, column=3).border = border
            written.add(key)
            r += 1

    # Restliche Attribute (nicht in Gruppen) am Ende
    remaining = {k: v for k, v in config.items() if k not in written}
    if remaining:
        c = ws.cell(row=r, column=1, value="Sonstige")
        c.font = Font(name="Arial", bold=True, size=10)
        c.fill = grp_fill
        c.border = border
        for col in [2, 3]:
            ws.cell(row=r, column=col).fill   = grp_fill
            ws.cell(row=r, column=col).border = border
        r += 1
        for key, val in sorted(remaining.items()):
            fill = val_fill if r % 2 == 0 else alt_fill
            ro   = key in _PLC_READONLY_ATTRS
            ws.cell(row=r, column=1, value=f"  {key}").fill   = fill
            ws.cell(row=r, column=1).font   = ro_font if ro else Font(name="Arial", size=10)
            ws.cell(row=r, column=1).border = border
            ws.cell(row=r, column=2, value=val or "").fill   = fill
            ws.cell(row=r, column=2).font   = Font(name="Arial", size=10)
            ws.cell(row=r, column=2).border = border
            ws.cell(row=r, column=3, value="Sonstige").fill   = fill
            ws.cell(row=r, column=3).font   = Font(name="Arial", size=10, color="808080")
            ws.cell(row=r, column=3).border = border
            r += 1

    ws.freeze_panes = "A2"
    wb.save(str(out))
    return {"status": "ok", "file": str(out), "device": device_name, "attributes": len(config)}


def export_hw_config(output_path=None):
    def _run():
        _sess.ensure_project()
        rows = []
        for device in _iter_all_devices(_sess.project):
            station = device.Name
            item_stack = [(item, 0) for item in device.DeviceItems]
            while item_stack:
                entry = item_stack.pop()
                di, depth = entry[0], entry[1]
                type_id = str(di.TypeIdentifier) if di.TypeIdentifier else ""
                ip = ""
                try:
                    v = di.GetAttribute("Address")
                    if v: ip = str(v)
                except Exception:
                    pass
                rows.append({
                    "Station":       station,
                    "Depth":         depth,
                    "Name":          str(di.Name),
                    "TypeIdentifier": type_id,
                    "Position":      str(getattr(di, "PositionNumber", "")),
                    "IP":            ip,
                })
                for sub in di.DeviceItems:
                    item_stack.append((sub, depth + 1))
        return rows

    rows = sta.run(_tia_call, _run)

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise TiaError("MISSING_DEPENDENCY", "openpyxl nicht installiert. pip install openpyxl", False)

    out = Path(output_path) if output_path else Path(_DEFAULT_EXPORT) / "hardware_config.xlsx"
    out.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Hardware"

    thin   = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    headers    = ["Station", "Komponente", "Bestellnummer", "Steckplatz", "IP-Adresse"]
    col_widths = [30, 42, 38, 12, 18]

    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font      = Font(name="Arial", bold=True, color="FFFFFF", size=10)
        c.fill      = PatternFill("solid", start_color="1F4E79")
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border    = border
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[1].height = 22

    prev_station = None
    for r, row in enumerate(rows, 2):
        s      = row["Station"]
        indent = "   " * row["Depth"]
        tid    = row["TypeIdentifier"].replace("OrderNumber:", "") if row["TypeIdentifier"] else ""
        bold   = row["Depth"] == 0
        bg     = "D6E4F0" if s != prev_station else ("EAF4FB" if row["Depth"] == 0 else "F5FBFF")
        fill   = PatternFill("solid", start_color=bg)
        vals   = [s if s != prev_station else "", indent + row["Name"], tid, row["Position"], row["IP"]]
        for col, val in enumerate(vals, 1):
            c           = ws.cell(row=r, column=col, value=val)
            c.font      = Font(name="Arial", size=10, bold=bold)
            c.fill      = fill
            c.border    = border
            c.alignment = Alignment(vertical="center")
        prev_station = s

    ws.freeze_panes  = "A2"
    ws.auto_filter.ref = f"A1:E{len(rows) + 1}"
    wb.save(str(out))
    return {"status": "ok", "file": str(out), "rows": len(rows)}


# ═══════════════════════════════════════════════════════════════════════════════
# HMI RUNTIME SETTINGS
# ═══════════════════════════════════════════════════════════════════════════════

def _rs_to_dict(obj, depth=0):
    """RuntimeSettings-Objekt rekursiv in dict umwandeln (max. 3 Ebenen)."""
    if obj is None or depth > 3:
        return None
    result = {}
    for ai in obj.GetAttributeInfos():
        name = str(ai.Name)
        val  = obj.GetAttribute(name)
        if val is None:
            result[name] = None
        elif hasattr(val, "GetAttributeInfos"):
            result[name] = _rs_to_dict(val, depth + 1)
        else:
            s = str(val)
            if s in ("True", "False"):
                result[name] = s == "True"
            else:
                result[name] = s
    return result


def _get_hmi_rs(device_name):
    """Unified HmiSoftware + RuntimeSettings-Objekt zurückgeben."""
    import Siemens.Engineering as eng
    sw_type = _get_sw_container_type()
    for device in _iter_all_devices(_sess.project):
        for item in device.DeviceItems:
            sw = _try_get_software(item, "", sw_type, eng)
            if sw and item.Name == device_name:
                rs = sw.GetAttribute("RuntimeSettings")
                if rs is None:
                    raise TiaError("NO_RUNTIME_SETTINGS",
                                   f"'{device_name}' hat kein RuntimeSettings-Attribut (nur Unified unterstützt).", False)
                return sw, rs
    raise TiaError("HMI_NOT_FOUND", f"HMI '{device_name}' nicht gefunden.", True)


def get_hmi_runtime_settings(device_name):
    def _run():
        _sess.ensure_project()
        _, rs = _get_hmi_rs(device_name)
        return {"status": "ok", "device": device_name, "settings": _rs_to_dict(rs)}
    return sta.run(_tia_call, _run)


def set_hmi_runtime_settings(device_name, settings: dict):
    """Setzt einfache (nicht-verschachtelte) RuntimeSettings-Werte."""
    def _run():
        _sess.ensure_project()
        _, rs = _get_hmi_rs(device_name)
        applied, skipped = [], []
        for key, val in settings.items():
            v = rs.GetAttribute(key)
            if v is not None and hasattr(v, "GetAttributeInfos"):
                skipped.append(key)
                continue
            rs.SetAttribute(key, val)
            applied.append(key)
        return {"status": "ok", "device": device_name, "applied": applied, "skipped_complex": skipped}
    return sta.run(_tia_call, _run)


def export_hmi_runtime_settings(device_name, output_path=None):
    def _run():
        _sess.ensure_project()
        _, rs = _get_hmi_rs(device_name)
        return _rs_to_dict(rs)

    settings = sta.run(_tia_call, _run)

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise TiaError("MISSING_DEPENDENCY", "openpyxl nicht installiert. pip install openpyxl", False)

    out = Path(output_path) if output_path else Path(_DEFAULT_EXPORT) / f"hmi_runtime_{device_name}.xlsx"
    out.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "RuntimeSettings"

    thin   = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    headers    = ["Einstellung", "Wert", "Gruppe"]
    col_widths = [42, 36, 30]
    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font      = Font(name="Arial", bold=True, color="FFFFFF", size=10)
        c.fill      = PatternFill("solid", start_color="1F4E79")
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border    = border
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[1].height = 22

    grp_fill  = PatternFill("solid", start_color="D6E4F0")
    val_fill  = PatternFill("solid", start_color="F5FBFF")
    alt_fill  = PatternFill("solid", start_color="EAF4FB")

    r = 2
    for key, val in settings.items():
        if isinstance(val, dict):
            c = ws.cell(row=r, column=1, value=key)
            c.font   = Font(name="Arial", bold=True, size=10)
            c.fill   = grp_fill
            c.border = border
            ws.cell(row=r, column=2).fill   = grp_fill
            ws.cell(row=r, column=2).border = border
            ws.cell(row=r, column=3).fill   = grp_fill
            ws.cell(row=r, column=3).border = border
            r += 1
            for sub_key, sub_val in val.items():
                fill = val_fill if r % 2 == 0 else alt_fill
                ws.cell(row=r, column=1, value=f"  {sub_key}").fill   = fill
                ws.cell(row=r, column=1).font   = Font(name="Arial", size=10)
                ws.cell(row=r, column=1).border = border
                ws.cell(row=r, column=1).alignment = Alignment(vertical="center")
                ws.cell(row=r, column=2, value=str(sub_val) if sub_val is not None else "").fill   = fill
                ws.cell(row=r, column=2).font   = Font(name="Arial", size=10)
                ws.cell(row=r, column=2).border = border
                ws.cell(row=r, column=3, value=key).fill   = fill
                ws.cell(row=r, column=3).font   = Font(name="Arial", size=10, color="808080")
                ws.cell(row=r, column=3).border = border
                r += 1
        else:
            fill = val_fill if r % 2 == 0 else alt_fill
            ws.cell(row=r, column=1, value=key).fill   = fill
            ws.cell(row=r, column=1).font   = Font(name="Arial", size=10, bold=True)
            ws.cell(row=r, column=1).border = border
            ws.cell(row=r, column=1).alignment = Alignment(vertical="center")
            ws.cell(row=r, column=2, value=str(val) if val is not None else "").fill   = fill
            ws.cell(row=r, column=2).font   = Font(name="Arial", size=10)
            ws.cell(row=r, column=2).border = border
            ws.cell(row=r, column=3).fill   = fill
            ws.cell(row=r, column=3).border = border
            r += 1

    ws.freeze_panes = "A2"
    wb.save(str(out))
    return {"status": "ok", "file": str(out), "device": device_name, "settings_count": len(settings)}


# ═══════════════════════════════════════════════════════════════════════════════
# HMI
# ═══════════════════════════════════════════════════════════════════════════════

def _get_hmi(device_name):
    import Siemens.Engineering as eng
    sw_type = _get_sw_container_type()
    all_names = []
    for device in _iter_all_devices(_sess.project):
        all_names.append(device.Name)
        if device.Name != device_name:
            continue
        for item in device.DeviceItems:
            sw = _try_get_software(item, "", sw_type, eng)
            if sw:
                # BUG-13 Fix: FullName prüfen, nicht nur __name__
                # IronPython gibt type.__name__ = "HmiSoftware" für BEIDE Typen zurück
                # aber type.__module__ unterscheidet: "HmiUnified" vs "Hmi"
                full = type(sw).__module__ + "." + type(sw).__name__
                if "HmiUnified" in full: return sw, "Unified"
                if "Hmi"        in full: return sw, "Advanced"
    raise TiaError("HMI_NOT_FOUND", f"HMI '{device_name}' nicht gefunden.", True,
                   {"available": all_names})

def _hmi_tag_tables(sw):
    """
    HMI Tag-Tabellen-Collection zurückgeben.
    Unified + Advanced neu:  sw.TagTables direkt
    Advanced HmiTarget:      sw.TagFolder.TagTables
    V19/V20 Unified alt:     sw.TagTableGroup.TagTables
    """
    if hasattr(sw, "TagTables"):
        return sw.TagTables
    if hasattr(sw, "TagFolder"):
        return sw.TagFolder.TagTables
    if hasattr(sw, "TagTableGroup"):
        return sw.TagTableGroup.TagTables
    return []

def _hmi_tag_tables_import(sw, fi, eng):
    """Import-Methode je nach API-Version."""
    if hasattr(sw, "TagTables"):
        sw.TagTables.Import(fi, eng.ImportOptions.Override)
    elif hasattr(sw, "TagFolder"):
        sw.TagFolder.TagTables.Import(fi, eng.ImportOptions.Override)
    elif hasattr(sw, "TagTableGroup"):
        sw.TagTableGroup.Import(fi, eng.ImportOptions.Override)
    else:
        raise TiaError("TAG_IMPORT_NOT_SUPPORTED", "Keine Import-Methode für HMI-Tags gefunden.", False)

def _hmi_screens(sw):
    """
    Alle HMI-Screens liefern — rekursiv aus Gruppen/Ordnern.
    Unified + Advanced neu:  Screens in sw.ScreenGroups (rekursiv)
    Advanced HmiTarget:      Screens in sw.ScreenFolder.Folders (rekursiv)
    Fallback:                sw.Screens direkt
    """
    screens = []

    def _collect_from_groups(collection):
        for grp in collection:
            for s in grp.Screens:
                screens.append(s)
            sub = getattr(grp, "ScreenGroups", None) or getattr(grp, "Folders", None)
            if sub:
                try: _collect_from_groups(sub)
                except Exception: pass

    def _collect_from_folder(folder):
        for s in folder.Screens:
            screens.append(s)
        for sub in folder.Folders:
            _collect_from_folder(sub)

    if hasattr(sw, "ScreenGroups"):
        _collect_from_groups(sw.ScreenGroups)
        if hasattr(sw, "Screens"):
            for s in sw.Screens:
                screens.append(s)
        return screens

    if hasattr(sw, "ScreenFolder"):
        _collect_from_folder(sw.ScreenFolder)
        return screens

    if hasattr(sw, "Screens"):
        return list(sw.Screens)
    if hasattr(sw, "ScreenCollection"):
        return list(sw.ScreenCollection)
    return []

def _hmi_screens_import(sw, fi, eng):
    """Screen-Import je nach API-Version."""
    if hasattr(sw, "Screens") and hasattr(sw.Screens, "Import"):
        sw.Screens.Import(fi, eng.ImportOptions.Override)
        return
    if hasattr(sw, "ScreenFolder"):
        sf = sw.ScreenFolder
        if hasattr(sf, "Screens") and hasattr(sf.Screens, "Import"):
            sf.Screens.Import(fi, eng.ImportOptions.Override)
            return
    if hasattr(sw, "ScreenCollection") and hasattr(sw.ScreenCollection, "Import"):
        sw.ScreenCollection.Import(fi, eng.ImportOptions.Override)
        return
    raise TiaError("SCREEN_IMPORT_NOT_SUPPORTED",
                   "Keine Import-Methode für HMI-Screens gefunden.", False)

def _hmi_screen_folders(sw):
    """
    Oberste Ordner-Collection für Screens zurückgeben.
    Unified + Advanced neu:  sw.ScreenGroups
    Advanced HmiTarget:      sw.ScreenFolder.Folders
    """
    if hasattr(sw, "ScreenGroups"):
        return sw.ScreenGroups, "ScreenGroups"
    if hasattr(sw, "ScreenFolders"):
        return sw.ScreenFolders, "ScreenFolders"
    if hasattr(sw, "ScreenFolder"):
        return sw.ScreenFolder.Folders, "Folders"
    return None, None

def _hmi_tag_folders(sw):
    """
    Oberste Ordner-Collection für Tag-Tabellen zurückgeben.
    Unified + Advanced neu:  sw.TagTableGroups
    Advanced HmiTarget:      sw.TagFolder.Folders
    Advanced älter:          sw.TagTableGroup (singular)
    """
    if hasattr(sw, "TagTableGroups"):
        return sw.TagTableGroups, "TagTableGroups"
    if hasattr(sw, "TagFolder"):
        return sw.TagFolder.Folders, "Folders"
    if hasattr(sw, "TagTableGroup"):
        return sw.TagTableGroup, "TagTableGroup"
    return None, None

def _ensure_folder(collection, name, sub_attr):
    """
    Gibt einen Ordner mit `name` aus `collection` zurück.
    Legt ihn an falls nicht vorhanden.
    sub_attr: Name der Unterordner-Collection am Ordner-Objekt (z.B. 'ScreenGroups').
    """
    for f in collection:
        if f.Name == name:
            return f
    return collection.Create(name)

def list_hmi_screens(device_name):
    def _run():
        _sess.ensure_project(); sw,ht = _get_hmi(device_name)
        screens = [{"name":s.Name,"width":getattr(s,"Width",None),
                    "height":getattr(s,"Height",None),
                    "items":s.ScreenItems.Count if hasattr(s,"ScreenItems") else None}
                   for s in _hmi_screens(sw)]
        return {"device":device_name,"hmi_type":ht,"screens":screens,"count":len(screens)}
    return sta.run(_tia_call, _run)

def list_hmi_tags(device_name, table_name=None):
    def _run():
        _sess.ensure_project(); sw,ht = _get_hmi(device_name); tags = []
        for table in _hmi_tag_tables(sw):
            if table_name and table.Name != table_name: continue
            for tag in table.Tags:
                tags.append({"name":tag.Name,"table":table.Name,
                    "type":str(getattr(tag,"DataTypeName","?")),
                    "high":getattr(tag,"HighLimit",None),
                    "low":getattr(tag,"LowLimit",None),
                    "archive":getattr(tag,"LoggingEnabled",None)})
        return {"device":device_name,"hmi_type":ht,"tags":tags,"count":len(tags)}
    return sta.run(_tia_call, _run)

def list_hmi_alarms(device_name):
    def _run():
        _sess.ensure_project(); sw,ht = _get_hmi(device_name); alarms = []
        for attr,kind in [("DiscreteAlarms","discrete"),("AnalogAlarms","analog"),("Alarms","unified")]:
            if hasattr(sw,attr):
                for a in getattr(sw,attr):
                    alarms.append({"name":a.Name,"type":kind,"class":str(getattr(a,"AlarmClass","?"))})
        return {"device":device_name,"hmi_type":ht,"alarms":alarms,"count":len(alarms)}
    return sta.run(_tia_call, _run)

def list_hmi_textlists(device_name):
    def _run():
        _sess.ensure_project(); sw,ht = _get_hmi(device_name); tls = []
        if hasattr(sw,"TextLists"):
            for tl in sw.TextLists:
                entries = [{"value":getattr(e,"Value",None),"text":str(getattr(e,"Text",""))}
                           for e in (tl.TextListEntries if hasattr(tl,"TextListEntries") else [])]
                tls.append({"name":tl.Name,"entries":entries})
        return {"device":device_name,"hmi_type":ht,"textlists":tls,"count":len(tls)}
    return sta.run(_tia_call, _run)

def _unified_screen_files(project_path):
    """
    Sucht Unified-Screen-Dateien (.hmiScreen) im TIA-Projektordner rekursiv.
    Gibt Dict zurück: {screen_stem: Path}
    """
    import glob as _glob
    proj_dir = Path(project_path).parent
    screens = {}
    for f in _glob.glob(str(proj_dir / "**" / "*.hmiScreen"), recursive=True):
        p = Path(f)
        screens[p.stem] = p
    return screens


def export_hmi_screen(device_name, screen_name, output_path):
    """
    Einzelnen HMI-Screen exportieren.
    Advanced: s.Export(FileInfo) — direkter API-Export.
    Unified:  Keine Export-API in V21 — .hmiScreen-Datei aus Projektordner kopieren.
    """
    def _run():
        _sess.ensure_project()
        import Siemens.Engineering as eng
        from System.IO import FileInfo
        sw, ht = _get_hmi(device_name)
        all_screens = _hmi_screens(sw)
        screen_names = [s.Name for s in all_screens]

        if screen_name not in screen_names:
            raise TiaError("SCREEN_NOT_FOUND", f"Screen '{screen_name}' nicht gefunden.", True,
                           {"available": screen_names})

        p = Path(output_path)
        p.parent.mkdir(parents=True, exist_ok=True)

        if ht == "Advanced":
            for s in all_screens:
                if s.Name == screen_name:
                    if p.exists(): p.unlink()
                    s.Export(FileInfo(str(p)), eng.ExportOptions.WithDefaults)
                    return {"status": "ok", "device": device_name, "hmi_type": ht,
                            "screen": screen_name, "output": str(p), "method": "api_export"}

        # Unified: Dateisystem
        proj_path = str(_sess.project.Path)
        screen_files = _unified_screen_files(proj_path)
        if screen_name not in screen_files:
            raise TiaError("SCREEN_FILE_NOT_FOUND",
                f"Screen '{screen_name}' nicht als .hmiScreen-Datei gefunden. "
                "Projekt speichern und erneut versuchen.",
                True, {"available_files": list(screen_files.keys())})
        import shutil as _shutil
        src = screen_files[screen_name]
        if p.exists(): p.unlink()
        _shutil.copy2(str(src), str(p))
        return {"status": "ok", "device": device_name, "hmi_type": ht,
                "screen": screen_name, "output": str(p),
                "method": "filesystem_copy", "source": str(src)}
    return sta.run(_tia_call, _run)

def export_hmi_tags(device_name, output_path=None):
    """Alle HMI Tag-Tabellen exportieren. output_path = Zielordner (Standard: C:\\tia-mcp\\export)."""
    def _run():
        _sess.ensure_project()
        import Siemens.Engineering as eng
        from System.IO import FileInfo, DirectoryInfo
        sw, ht = _get_hmi(device_name)
        out_dir = _export_dir(output_path)

        exported = []
        tables = _hmi_tag_tables(sw)
        for table in tables:
            safe_name = table.Name.replace(" ", "_").replace("/", "_")
            if safe_name == "Default_tag_table":
                continue  # Default-Tabelle überspringen

            # Advanced HmiTarget / ältere API: table.Export(FileInfo)
            if hasattr(table, "Export"):
                xml_file = out_dir / f"hmi_tags_{safe_name}.xml"
                if xml_file.exists(): xml_file.unlink()
                table.Export(FileInfo(str(xml_file)), eng.ExportOptions.WithDefaults)
                exported.append(str(xml_file))

            # Unified V21: table.Tags.Export(DirectoryInfo) — Workaround
            elif hasattr(table, "Tags") and hasattr(table.Tags, "Export"):
                tag_dir = out_dir / f"hmi_tags_{safe_name}"
                tag_dir.mkdir(parents=True, exist_ok=True)
                table.Tags.Export(DirectoryInfo(str(tag_dir)))
                exported.append(str(tag_dir))

        # Fallback: TagTableGroups — Tabellen in Untergruppen
        if not exported and hasattr(sw, "TagTableGroups"):
            for grp in sw.TagTableGroups:
                for table in grp.TagTables:
                    safe_name = table.Name.replace(" ", "_").replace("/", "_")
                    if hasattr(table, "Tags") and hasattr(table.Tags, "Export"):
                        tag_dir = out_dir / f"hmi_tags_{safe_name}"
                        tag_dir.mkdir(parents=True, exist_ok=True)
                        table.Tags.Export(DirectoryInfo(str(tag_dir)))
                        exported.append(str(tag_dir))

        if not exported:
            raise TiaError("NO_TAG_TABLES",
                f"HMI '{device_name}' hat keine exportierbaren Tag-Tabellen.", True)

        return {"status": "ok", "device": device_name, "hmi_type": ht,
                "exported": exported, "count": len(exported)}
    return sta.run(_tia_call, _run)

# ═══════════════════════════════════════════════════════════════════════════════
# BIBLIOTHEK
# ═══════════════════════════════════════════════════════════════════════════════

def _lib_name(lib):
    """Bibliotheksname sicher auslesen — V21 hat kein .Name auf ProjectLibrary."""
    for attr in ("Name", "Caption", "RootGroup"):
        try:
            v = getattr(lib, attr, None)
            if v is not None:
                return str(v)
        except Exception:
            pass
    return "ProjectLibrary"

def _global_libraries():
    """GlobalLibraries — in V21 am Portal, in V19/V20 am Projekt."""
    for src in (_sess.portal, _sess.project):
        if src is None:
            continue
        try:
            return list(src.GlobalLibraries)
        except Exception:
            pass
    return []

def _find_lib(name):
    p = _sess.project
    pl_name = _lib_name(p.ProjectLibrary)
    if pl_name == name:
        return p.ProjectLibrary, "project"
    for gl in _global_libraries():
        if _lib_name(gl) == name:
            return gl, "global"
    raise TiaError("LIBRARY_NOT_FOUND", f"Bibliothek '{name}' nicht gefunden.", True,
                   {"available": [pl_name] + [_lib_name(gl) for gl in _global_libraries()]})

def _collect_types(folder, result, path=""):
    for t in folder.Types:
        fp = f"{path}/{t.Name}".lstrip("/"); dv = None; versions = []
        try:
            for v in t.Versions:
                try: is_def = t.DefaultVersion and str(v.VersionNumber)==str(t.DefaultVersion.VersionNumber)
                except: is_def = False
                versions.append({"version":str(v.VersionNumber),"state":str(v.State),"is_default":is_def})
                if is_def: dv = str(v.VersionNumber)
        except Exception as e: versions = [{"error":str(e)}]
        result.append({"name":t.Name,"path":fp,"versions":versions,"default_version":dv})
    for sub in folder.Folders:
        _collect_types(sub, result, f"{path}/{sub.Name}".lstrip("/"))

def list_libraries():
    def _run():
        _sess.ensure_project(); p = _sess.project
        pl = p.ProjectLibrary
        pl_name = _lib_name(pl)
        libs = [{"name": pl_name, "scope": "project",
                 "types":  pl.TypeFolder.Types.Count,
                 "copies": pl.MasterCopyFolder.MasterCopies.Count}]
        for gl in _global_libraries():
            libs.append({"name": _lib_name(gl), "scope": "global", "path": str(gl.Path),
                         "types":  gl.TypeFolder.Types.Count,
                         "copies": gl.MasterCopyFolder.MasterCopies.Count})
        return {"libraries": libs, "count": len(libs)}
    return sta.run(_tia_call, _run)

def list_library_types(library_name):
    def _run():
        _sess.ensure_project(); lib,scope = _find_lib(library_name)
        types_list = []; _collect_types(lib.TypeFolder, types_list)
        return {"library":library_name,"scope":scope,"types":types_list,"count":len(types_list)}
    return sta.run(_tia_call, _run)

def list_master_copies(library_name):
    def _run():
        _sess.ensure_project(); lib,scope = _find_lib(library_name)
        def _col(folder, path=""):
            r = [{"name":mc.Name,"path":f"{path}/{mc.Name}".lstrip("/")} for mc in folder.MasterCopies]
            for sub in folder.Folders: r.extend(_col(sub,f"{path}/{sub.Name}".lstrip("/")))
            return r
        copies = _col(lib.MasterCopyFolder)
        return {"library":library_name,"scope":scope,"master_copies":copies,"count":len(copies)}
    return sta.run(_tia_call, _run)

def get_library_type_versions(library_name, type_name):
    def _run():
        _sess.ensure_project(); lib,scope = _find_lib(library_name)
        types_list = []; _collect_types(lib.TypeFolder, types_list)
        for t in types_list:
            if t["name"] == type_name:
                return {"library":library_name,"type":type_name,
                        "versions":t["versions"],"default_version":t["default_version"]}
        raise TiaError("TYPE_NOT_FOUND",f"Typ '{type_name}' nicht gefunden.",True,
                       {"available":[t["name"] for t in types_list]})
    return sta.run(_tia_call, _run)

# ═══════════════════════════════════════════════════════════════════════════════
# EXECUTOR
# ═══════════════════════════════════════════════════════════════════════════════

_BLOCKED_WRITE  = ["save","delete","remove","create","import","compile",
                   "download","export","update","set","add","insert","copy","move","rename"]
_BLOCKED_ALWAYS = ["exec","eval","open","os.","sys.","subprocess","shutil","__import__","builtins"]
_SAFE_BUILTINS  = {
    "len":len,"str":str,"int":int,"float":float,"bool":bool,"list":list,"dict":dict,
    "tuple":tuple,"set":set,"print":print,"range":range,"enumerate":enumerate,
    "zip":zip,"map":map,"filter":filter,"sorted":sorted,"hasattr":hasattr,
    "getattr":getattr,"isinstance":isinstance,"type":type,
    "None":None,"True":True,"False":False,
}

def _iter_all_devices(project):
    """
    Liefert alle Devices aus project.Devices UND allen project.DeviceGroups (rekursiv).
    Nötig weil TIA-Projekte mit Gerätegruppen (Ordner) die Devices nicht auf Top-Level haben.
    """
    # Top-Level
    for device in project.Devices:
        yield device
    # DeviceGroups rekursiv (Stack statt Rekursion)
    try:
        group_stack = list(project.DeviceGroups)
        while group_stack:
            grp = group_stack.pop()
            for device in grp.Devices:
                yield device
            try:
                for sub in grp.Groups:
                    group_stack.append(sub)
            except Exception:
                pass
    except Exception:
        pass

def _get_sw_container_type():
    """
    SoftwareContainer-Typ laden.
    V21: Siemens.Engineering.HW.Features.SoftwareContainer (Base-Assembly)
    V19/V20: Siemens.Engineering.SW.SoftwareContainer
    """
    import clr
    candidates = [
        # V21 — exakter Pfad aus Reflektion ermittelt
        "Siemens.Engineering.HW.Features.SoftwareContainer",
        # V19/V20 Fallback
        "Siemens.Engineering.SW.SoftwareContainer",
    ]
    for full_name in candidates:
        parts = full_name.rsplit(".", 1)
        ns, cls = parts[0], parts[1]
        try:
            mod = __import__(ns, fromlist=[cls])
            t = getattr(mod, cls, None)
            if t:
                _log("session").info(f"SoftwareContainer: {full_name}")
                return t
        except Exception: pass
    # Letzter Ausweg: ueber Reflektion auf geladene Assemblies
    try:
        import System
        for asm in System.AppDomain.CurrentDomain.GetAssemblies():
            t = asm.GetType("Siemens.Engineering.HW.Features.SoftwareContainer")
            if t:
                return t
    except Exception: pass
    return None

_SW_CONTAINER_TYPE = None

def _find_sw(project, name, hint, eng):
    """
    Software-Objekt suchen — iterativ, alle Device- und DeviceItem-Ebenen.
    name: DeviceItem-Name (z.B. 'PLC_1') oder leer fuer erstes passendes.
    hint: 'PlcSoftware', 'Hmi', 'Unified' oder leer.
    """
    global _SW_CONTAINER_TYPE
    if _SW_CONTAINER_TYPE is None:
        _SW_CONTAINER_TYPE = _get_sw_container_type()

    sw_type = _SW_CONTAINER_TYPE

    for device in _iter_all_devices(project):
        stack = list(device.DeviceItems)
        while stack:
            item = stack.pop()
            name_match = (not name) or (item.Name == name) or (device.Name == name)
            if name_match:
                # Alle bekannten Wege versuchen
                sw = _try_get_software(item, hint, sw_type, eng)
                if sw: return sw
            try:
                for sub in item.DeviceItems: stack.append(sub)
            except Exception: pass
    return None

def _try_get_software(item, hint, sw_type, eng):
    """GetService mit allen bekannten SoftwareContainer-Typen versuchen."""
    type_candidates = []
    if sw_type: type_candidates.append(sw_type)
    # Fallbacks
    for tc in [
        lambda: eng.SW.SoftwareContainer,
        lambda: eng.HW.SoftwareContainer,
    ]:
        try: type_candidates.append(tc())
        except Exception: pass

    for t in type_candidates:
        try:
            swc = item.GetService[t]()
            if swc:
                sw = swc.Software
                tn = type(sw).__name__
                if not hint or hint.lower() in tn.lower():
                    return sw
        except Exception: pass
    return None

def _to_json(obj):
    if obj is None: return None
    if isinstance(obj,(bool,int,float,str)): return obj
    if isinstance(obj,(list,tuple)): return [_to_json(i) for i in obj]
    if isinstance(obj,dict): return {str(k):_to_json(v) for k,v in obj.items()}
    try: return str(obj)
    except: return f"<{type(obj).__name__}>"


def _mltext(obj, lang="de-DE"):
    """MultilingualText-Objekt sicher als String auslesen (BUG-15)."""
    if obj is None:
        return ""
    if isinstance(obj, str):
        return obj
    try:
        items = obj.Items
        for item in items:
            if str(getattr(item, "Language", "")).startswith(lang[:2]):
                return str(item.Text) if item.Text else ""
        for item in items:
            t = str(item.Text) if item.Text else ""
            if t:
                return t
    except Exception:
        pass
    return ""

def execute_openness(code, mode="read"):
    cl = code.lower()
    for kw in _BLOCKED_ALWAYS:
        if kw in cl:
            raise TiaError("CODE_BLOCKED",f"'{kw}' nicht erlaubt.",False,{"kw":kw})
    if mode == "read":
        for kw in _BLOCKED_WRITE:
            if f".{kw}(" in cl:
                raise TiaError("WRITE_BLOCKED",
                    f"'.{kw}()' im read-Modus gesperrt. mode='write' verwenden.",True,{"call":f".{kw}()"})
    _log("executor").info(f"execute mode={mode} {len(code)}ch")

    def _run():
        _sess.ensure_portal()
        import Siemens.Engineering as eng
        import Siemens.Engineering.SW as eng_sw
        hmi_ns = unified_ns = None
        try: import Siemens.Engineering.Hmi as hmi_ns
        except Exception: pass
        try: import Siemens.Engineering.HmiUnified as unified_ns
        except Exception: pass
        # SoftwareContainer-Typ fuer direkten Zugriff im Code
        sw_container_type = _get_sw_container_type()

        ctx = {
            "portal":   _sess.portal,
            "project":  _sess.project,
            "eng":      eng,
            "sw":       eng_sw,
            "hmi":      hmi_ns,
            "unified":  unified_ns,
            "result":   None,
            # SoftwareContainer direkt nutzbar:
            # swc = item.GetService[SoftwareContainer]()
            "SoftwareContainer": sw_container_type,
            "safe_str":      lambda o: str(o) if o is not None else None,
            "collect":       lambda c: list(c),
            "find_software": lambda n, h="": _find_sw(_sess.project, n, h, eng),
            # Alle Devices inkl. DeviceGroups iterieren:
            # for dev in iter_devices(): ...
            "iter_devices":  lambda: _iter_all_devices(_sess.project),
        }
        exec(textwrap.dedent(code), {"__builtins__":_SAFE_BUILTINS}, ctx)
        return {"status":"ok","mode":mode,"result":_to_json(ctx.get("result"))}
    return sta.run(_tia_call, _run)

# ── Setup / Teardown ──────────────────────────────────────────────────────────
def setup(log_dir="C:/tia-mcp/logs"):
    _setup_logging(log_dir); sta.start()

def teardown():
    sta.stop()

# ═══════════════════════════════════════════════════════════════════════════════
# PLC EXPORT / IMPORT
# ═══════════════════════════════════════════════════════════════════════════════

_DEFAULT_EXPORT = r"C:\tia-mcp\export"

def _export_dir(path=None):
    """Exportpfad — Standard oder Override."""
    p = Path(path or _DEFAULT_EXPORT)
    p.mkdir(parents=True, exist_ok=True)
    return p

def _find_block(plc_sw, block_name):
    """Baustein iterativ suchen — alle Gruppen."""
    stack = [plc_sw.BlockGroup]
    while stack:
        group = stack.pop()
        for block in group.Blocks:
            if block.Name == block_name:
                return block
        for sub in group.Groups:
            stack.append(sub)
    return None

def list_plc_blocks(device_name, group=None):
    """
    Alle PLC-Bausteine auflisten.
    Durchsucht BlockGroup rekursiv inkl. Untergruppen.
    group: optionaler Gruppenname für Filterung (z.B. "Ventile").
    Rückgabe je Baustein: name, type (OB/FC/FB/DB), language, number, group, is_consistent.
    """
    def _run():
        _sess.ensure_project()
        import Siemens.Engineering as eng
        plc = _find_sw(_sess.project, device_name, "PlcSoftware", eng)
        if not plc:
            raise TiaError("PLC_NOT_FOUND", f"PLC '{device_name}' nicht gefunden.", True)

        blocks = []

        def _collect(bg, path):
            for b in bg.Blocks:
                btype = b.GetType().Name  # OB, FC, FB, DB, PlcStruct etc.
                lang  = str(getattr(b, "ProgrammingLanguage", "?"))
                blocks.append({
                    "name":          b.Name,
                    "type":          btype,
                    "language":      lang,
                    "number":        getattr(b, "Number", None),
                    "group":         path,
                    "is_consistent": getattr(b, "IsConsistent", None),
                })
            for g in bg.Groups:
                gpath = (path + "/" if path else "") + g.Name
                _collect(g, gpath)

        _collect(plc.BlockGroup, "")

        # Filtern nach Gruppe wenn angegeben
        if group:
            blocks = [b for b in blocks if b["group"] == group or b["group"].startswith(group + "/")]

        return {
            "device":  device_name,
            "count":   len(blocks),
            "blocks":  blocks,
        }
    return sta.run(_tia_call, _run)


def list_plc_tag_tables(device_name):
    """
    Alle PLC-Tag-Tabellen auflisten inkl. Untergruppen.
    Rückgabe je Tabelle: name, group, tag_count.
    """
    def _run():
        _sess.ensure_project()
        import Siemens.Engineering as eng
        plc = _find_sw(_sess.project, device_name, "PlcSoftware", eng)
        if not plc:
            raise TiaError("PLC_NOT_FOUND", f"PLC '{device_name}' nicht gefunden.", True)

        tables = []

        def _collect(ttg, path):
            for tt in ttg.TagTables:
                tables.append({
                    "name":      tt.Name,
                    "group":     path,
                    "tag_count": tt.Tags.Count,
                })
            for g in ttg.Groups:
                gpath = (path + "/" if path else "") + g.Name
                _collect(g, gpath)

        _collect(plc.TagTableGroup, "")

        return {
            "device": device_name,
            "count":  len(tables),
            "tables": tables,
        }
    return sta.run(_tia_call, _run)


def list_plc_tags(device_name, table_name):
    """
    Alle Tags einer PLC-Tag-Tabelle auflisten.
    Rückgabe je Tag: name, data_type, logical_address, comment, visible.
    """
    def _run():
        _sess.ensure_project()
        import Siemens.Engineering as eng
        plc = _find_sw(_sess.project, device_name, "PlcSoftware", eng)
        if not plc:
            raise TiaError("PLC_NOT_FOUND", f"PLC '{device_name}' nicht gefunden.", True)

        # Tag-Tabelle suchen (Top-Level + Gruppen)
        found_table = None

        def _find(ttg):
            for tt in ttg.TagTables:
                if tt.Name == table_name:
                    return tt
            for g in ttg.Groups:
                result = _find(g)
                if result:
                    return result
            return None

        found_table = _find(plc.TagTableGroup)
        if not found_table:
            # Available-Liste aufbauen
            all_tables = []
            def _list(ttg):
                for tt in ttg.TagTables:
                    all_tables.append(tt.Name)
                for g in ttg.Groups:
                    _list(g)
            _list(plc.TagTableGroup)
            raise TiaError("TABLE_NOT_FOUND",
                f"Tag-Tabelle '{table_name}' nicht gefunden.", True,
                {"available": all_tables})

        tags = []
        for tag in found_table.Tags:
            tags.append({
                "name":            tag.Name,
                "data_type":       str(getattr(tag, "DataTypeName", "?")),
                "logical_address": str(getattr(tag, "LogicalAddress", "")),
                "comment":         _mltext(getattr(tag, "Comment", None)),
                "visible":         getattr(tag, "Visible", None),
            })

        return {
            "device":     device_name,
            "table":      table_name,
            "tag_count":  len(tags),
            "tags":       tags,
        }
    return sta.run(_tia_call, _run)


def list_plc_udts(device_name):
    """
    Alle PLC-UDTs (Strukturen) auflisten inkl. Untergruppen.
    Rückgabe je UDT: name, group, is_consistent, modified_date.
    """
    def _run():
        _sess.ensure_project()
        import Siemens.Engineering as eng
        plc = _find_sw(_sess.project, device_name, "PlcSoftware", eng)
        if not plc:
            raise TiaError("PLC_NOT_FOUND", f"PLC '{device_name}' nicht gefunden.", True)

        udts = []

        def _collect(tg, path):
            for udt in tg.Types:
                udts.append({
                    "name":          udt.Name,
                    "type":          udt.GetType().Name,  # PlcStruct, PlcTypedeF...
                    "group":         path,
                    "is_consistent": getattr(udt, "IsConsistent", None),
                    "modified_date": str(getattr(udt, "ModifiedDate", "") or ""),
                })
            for g in tg.Groups:
                gpath = (path + "/" if path else "") + g.Name
                _collect(g, gpath)

        _collect(plc.TypeGroup, "")

        return {
            "device": device_name,
            "count":  len(udts),
            "udts":   udts,
        }
    return sta.run(_tia_call, _run)


def export_plc_block(device_name, block_name, output_path=None):
    def _run():
        _sess.ensure_project()
        import Siemens.Engineering as eng
        from System.IO import FileInfo
        sw_type = _get_sw_container_type()
        plc = _find_sw(_sess.project, device_name, "PlcSoftware", eng)
        if not plc:
            raise TiaError("PLC_NOT_FOUND", f"PLC '{device_name}' nicht gefunden.", True)
        block = _find_block(plc, block_name)
        if not block:
            avail = [b.Name for b in plc.BlockGroup.Blocks]
            raise TiaError("BLOCK_NOT_FOUND", f"Baustein '{block_name}' nicht gefunden.", True,
                           {"available": avail})
        out_dir = _export_dir(output_path)
        xml_file = out_dir / f"{block_name}.xml"
        if xml_file.exists(): xml_file.unlink()   # TIA überschreibt nicht
        block.Export(FileInfo(str(xml_file)), eng.ExportOptions.WithDefaults)
        _log("plc").info(f"Exportiert: {block_name} → {xml_file}")
        return {"status": "ok", "block": block_name, "type": type(block).__name__,
                "xml_path": str(xml_file)}
    return sta.run(_tia_call, _run)

def import_plc_block(device_name, file_path):
    def _run():
        _sess.ensure_project()
        import Siemens.Engineering as eng
        from System.IO import FileInfo
        plc = _find_sw(_sess.project, device_name, "PlcSoftware", eng)
        if not plc:
            raise TiaError("PLC_NOT_FOUND", f"PLC '{device_name}' nicht gefunden.", True)
        fi = FileInfo(file_path)
        if not fi.Exists:
            raise TiaError("FILE_NOT_FOUND", f"Datei nicht gefunden: {file_path}", True)
        result = plc.BlockGroup.Blocks.Import(fi, eng.ImportOptions.Override)
        _log("plc").info(f"Importiert: {file_path}")
        return {"status": "ok", "imported_from": file_path,
                "blocks": [str(b) for b in result] if result else []}
    return sta.run(_tia_call, _run)

def get_plc_block_source(device_name, block_name, output_path=None):
    """
    Exportiert Baustein und liest den Quellcode aus.
    SCL-Blöcke: gibt lesbaren SCL-Code zurück.
    LAD/FBD: gibt XML zurück.
    """
    def _run():
        _sess.ensure_project()
        import Siemens.Engineering as eng
        from System.IO import FileInfo
        plc = _find_sw(_sess.project, device_name, "PlcSoftware", eng)
        if not plc:
            raise TiaError("PLC_NOT_FOUND", f"PLC '{device_name}' nicht gefunden.", True)
        block = _find_block(plc, block_name)
        if not block:
            raise TiaError("BLOCK_NOT_FOUND", f"Baustein '{block_name}' nicht gefunden.", True)

        out_dir  = _export_dir(output_path)
        xml_file = out_dir / f"{block_name}.xml"
        if xml_file.exists(): xml_file.unlink()   # TIA überschreibt nicht
        block.Export(FileInfo(str(xml_file)), eng.ExportOptions.WithDefaults)

        # XML lesen
        xml_content = xml_file.read_text(encoding="utf-8")
        block_type  = type(block).__name__
        language    = getattr(block, "ProgrammingLanguage", None)
        lang_str    = str(language) if language else "Unknown"

        # SCL-Quellcode extrahieren
        scl_source  = None
        scl_file    = None
        if "SCL" in lang_str.upper() or "StructuredControlLanguage" in lang_str:
            scl_source = _extract_scl(xml_content)
            if scl_source:
                scl_file = out_dir / f"{block_name}.scl"
                scl_file.write_text(scl_source, encoding="utf-8")

        return {
            "status":    "ok",
            "block":     block_name,
            "type":      block_type,
            "language":  lang_str,
            "xml_path":  str(xml_file),
            "scl_path":  str(scl_file) if scl_file else None,
            "scl_source": scl_source,
            "xml_size_kb": round(len(xml_content) / 1024, 1)
        }
    return sta.run(_tia_call, _run)

def _extract_scl(xml_content: str) -> str | None:
    """
    SCL-Quellcode aus TIA-XML extrahieren.
    V19/V20: <StructuredText>CODE</StructuredText> oder <Body>CODE</Body>
    V21:     <StructuredText xmlns="...v4"><Token Text="..."/><Token Text="..."/>...</StructuredText>
    """
    import re
    # V21: Token-basiertes Format (Schema v4)
    m = re.search(r"<StructuredText[^>]*>(.*?)</StructuredText>", xml_content, re.DOTALL)
    if m:
        inner = m.group(1).strip()
        if inner:
            # Token-Elemente zusammensetzen
            tokens = re.findall(r'<Token[^>]+Text="([^"]*)"', inner)
            if tokens:
                return "".join(tokens)
            # Kein Token — CDATA oder reiner Text
            src = re.sub(r"<!\[CDATA\[(.*?)\]\]>", r"\1", inner, flags=re.DOTALL).strip()
            if src:
                return src
    # Ältere Formate
    for tag in ["Body", "SourceText"]:
        m = re.search(rf"<{tag}>(.*?)</{tag}>", xml_content, re.DOTALL)
        if m:
            src = m.group(1).strip()
            src = re.sub(r"<!\[CDATA\[(.*?)\]\]>", r"\1", src, flags=re.DOTALL)
            if src:
                return src
    return None

# ═══════════════════════════════════════════════════════════════════════════════
# PLC TAG-TABELLEN EXPORT / IMPORT
# ═══════════════════════════════════════════════════════════════════════════════

def export_plc_tagtable(device_name, table_name, output_path=None):
    def _run():
        _sess.ensure_project()
        import Siemens.Engineering as eng
        from System.IO import FileInfo
        plc = _find_sw(_sess.project, device_name, "PlcSoftware", eng)
        if not plc:
            raise TiaError("PLC_NOT_FOUND", f"PLC '{device_name}' nicht gefunden.", True)
        for table in plc.TagTableGroup.TagTables:
            if table.Name == table_name:
                out_dir  = _export_dir(output_path)
                xml_file = out_dir / f"plc_tags_{table_name}.xml"
                if xml_file.exists(): xml_file.unlink()
                table.Export(FileInfo(str(xml_file)), eng.ExportOptions.WithDefaults)
                return {"status": "ok", "table": table_name, "xml_path": str(xml_file)}
        raise TiaError("TABLE_NOT_FOUND", f"Tag-Tabelle '{table_name}' nicht gefunden.", True)
    return sta.run(_tia_call, _run)

def import_plc_tagtable(device_name, file_path):
    def _run():
        _sess.ensure_project()
        import Siemens.Engineering as eng
        from System.IO import FileInfo
        plc = _find_sw(_sess.project, device_name, "PlcSoftware", eng)
        if not plc:
            raise TiaError("PLC_NOT_FOUND", f"PLC '{device_name}' nicht gefunden.", True)
        fi = FileInfo(file_path)
        if not fi.Exists:
            raise TiaError("FILE_NOT_FOUND", f"Datei nicht gefunden: {file_path}", True)
        plc.TagTableGroup.TagTables.Import(fi, eng.ImportOptions.Override)
        return {"status": "ok", "imported_from": file_path}
    return sta.run(_tia_call, _run)

# ═══════════════════════════════════════════════════════════════════════════════
# HMI EXPORT / IMPORT
# ═══════════════════════════════════════════════════════════════════════════════

def export_hmi_tagtable(device_name, table_name, output_path=None):
    def _run():
        _sess.ensure_project()
        import Siemens.Engineering as eng
        from System.IO import FileInfo, DirectoryInfo
        sw, ht = _get_hmi(device_name)
        for table in _hmi_tag_tables(sw):
            if table.Name == table_name:
                # V19/V20: Export() direkt auf Tabelle
                if hasattr(table, "Export") and callable(getattr(table, "Export")):
                    out_dir  = _export_dir(output_path)
                    xml_file = out_dir / f"hmi_tags_{table_name}.xml"
                    if xml_file.exists(): xml_file.unlink()
                    table.Export(FileInfo(str(xml_file)), eng.ExportOptions.WithDefaults)
                    return {"status": "ok", "device": device_name, "hmi_type": ht,
                            "table": table_name, "xml_path": str(xml_file)}
                # V21 WinCC Advanced/Unified: Tags.Export(DirectoryInfo)
                if hasattr(table, "Tags") and hasattr(table.Tags, "Export"):
                    out_dir = _export_dir(output_path) / f"hmi_tags_{table_name}"
                    out_dir.mkdir(parents=True, exist_ok=True)
                    files = list(table.Tags.Export(DirectoryInfo(str(out_dir))))
                    exported = [str(f) for f in files]
                    return {"status": "ok", "device": device_name, "hmi_type": ht,
                            "table": table_name, "exported": exported,
                            "export_dir": str(out_dir)}
                raise TiaError("TAG_EXPORT_NOT_SUPPORTED",
                    f"HMI Tag-Tabelle '{table_name}' gefunden, aber kein Export verfügbar "
                    f"({ht}). Workaround: list_hmi_tags + write_import_file.", False,
                    {"table": table_name, "hmi_type": ht})
        available = [t.Name for t in _hmi_tag_tables(sw)]
        raise TiaError("TABLE_NOT_FOUND", f"HMI Tag-Tabelle '{table_name}' nicht gefunden.", True,
                       {"available": available})
    return sta.run(_tia_call, _run)

def import_hmi_tagtable(device_name, file_path):
    def _run():
        _sess.ensure_project()
        import Siemens.Engineering as eng
        from System.IO import FileInfo, DirectoryInfo
        sw, ht = _get_hmi(device_name)
        p = Path(file_path)
        if not p.exists():
            raise TiaError("FILE_NOT_FOUND", f"Datei/Ordner nicht gefunden: {file_path}", True)
        # Ordner-Import (V21 Tags.Import(DirectoryInfo))
        if p.is_dir():
            for table in _hmi_tag_tables(sw):
                if hasattr(table, "Tags") and hasattr(table.Tags, "Import"):
                    table.Tags.Import(DirectoryInfo(str(p)))
                    return {"status": "ok", "device": device_name, "imported_from": file_path}
        # Datei-Import (V19/V20)
        _hmi_tag_tables_import(sw, FileInfo(str(p)), eng)
        return {"status": "ok", "device": device_name, "imported_from": file_path}
    return sta.run(_tia_call, _run)


def import_hmi_tags(device_name, file_path):
    """Alle HMI-Tags aus Datei oder Ordner importieren."""
    def _run():
        _sess.ensure_project()
        import Siemens.Engineering as eng
        from System.IO import FileInfo, DirectoryInfo
        sw, ht = _get_hmi(device_name)
        p = Path(file_path)
        if not p.exists():
            raise TiaError("FILE_NOT_FOUND", f"Datei/Ordner nicht gefunden: {file_path}", True)
        if p.is_dir():
            for table in _hmi_tag_tables(sw):
                if hasattr(table, "Tags") and hasattr(table.Tags, "Import"):
                    table.Tags.Import(DirectoryInfo(str(p)))
            return {"status": "ok", "device": device_name, "imported_from": file_path}
        _hmi_tag_tables_import(sw, FileInfo(str(p)), eng)
        return {"status": "ok", "device": device_name, "imported_from": file_path}
    return sta.run(_tia_call, _run)


def export_hmi_screens_all(device_name, output_path=None):
    """
    Alle HMI-Screens exportieren.
    Advanced: s.Export(FileInfo). Unified: .hmiScreen-Dateien aus Projektordner kopieren.
    """
    def _run():
        _sess.ensure_project()
        import Siemens.Engineering as eng
        from System.IO import FileInfo
        sw, ht = _get_hmi(device_name)
        out_dir = (_export_dir(None) / f"hmi_screens_{device_name}") \
                  if not output_path else Path(output_path)
        out_dir.mkdir(parents=True, exist_ok=True)
        all_screens = _hmi_screens(sw)
        if not all_screens:
            raise TiaError("NO_SCREENS", f"HMI '{device_name}' hat keine Screens.", True)
        exported = []
        if ht == "Advanced":
            for s in all_screens:
                safe_name = s.Name.replace(" ", "_").replace("/", "_")
                xml_file = out_dir / f"{safe_name}.xml"
                if xml_file.exists(): xml_file.unlink()
                try:
                    s.Export(FileInfo(str(xml_file)), eng.ExportOptions.WithDefaults)
                    exported.append({"screen": s.Name, "path": str(xml_file), "method": "api_export"})
                except Exception as e:
                    exported.append({"screen": s.Name, "error": str(e)})
        else:
            import shutil as _shutil
            proj_path = str(_sess.project.Path)
            screen_files = _unified_screen_files(proj_path)
            for s in all_screens:
                name = s.Name
                if name in screen_files:
                    src = screen_files[name]
                    dst = out_dir / src.name
                    if dst.exists(): dst.unlink()
                    _shutil.copy2(str(src), str(dst))
                    exported.append({"screen": name, "path": str(dst), "method": "filesystem_copy"})
                else:
                    exported.append({"screen": name,
                                     "error": "Keine .hmiScreen-Datei — Projekt gespeichert?"})
        ok_count = len([e for e in exported if "path" in e])
        return {"status": "ok", "device": device_name, "hmi_type": ht,
                "exported": exported, "count": ok_count}
    return sta.run(_tia_call, _run)

def export_hmi_scripts(device_name, output_path=None):
    """
    HMI-Scripts exportieren.
    Unified:            sw.Scripts.Export(DirectoryInfo)
    Advanced HmiTarget: sw.VBScriptFolder.VBScripts einzeln exportieren
    """
    def _run():
        _sess.ensure_project()
        import Siemens.Engineering as eng
        from System.IO import FileInfo, DirectoryInfo
        sw, ht = _get_hmi(device_name)
        out_dir = (_export_dir(None) / f"hmi_scripts_{device_name}") \
                  if not output_path else Path(output_path)
        out_dir.mkdir(parents=True, exist_ok=True)
        exported = []

        # Unified: sw.Scripts
        scripts = getattr(sw, "Scripts", None)
        if scripts is not None:
            if hasattr(scripts, "Export"):
                files = list(scripts.Export(DirectoryInfo(str(out_dir))))
                exported = [str(f) for f in files]
            else:
                for sc in scripts:
                    safe_name = sc.Name.replace(" ", "_").replace("/", "_")
                    if hasattr(sc, "Export"):
                        xml_file = out_dir / f"{safe_name}.xml"
                        if xml_file.exists(): xml_file.unlink()
                        sc.Export(FileInfo(str(xml_file)), eng.ExportOptions.WithDefaults)
                        exported.append(str(xml_file))

        # Advanced HmiTarget: sw.VBScriptFolder
        if not exported and hasattr(sw, "VBScriptFolder"):
            vbf = sw.VBScriptFolder
            for sc in vbf.VBScripts:
                safe_name = sc.Name.replace(" ", "_").replace("/", "_")
                if hasattr(sc, "Export"):
                    xml_file = out_dir / f"{safe_name}.xml"
                    if xml_file.exists(): xml_file.unlink()
                    sc.Export(FileInfo(str(xml_file)), eng.ExportOptions.WithDefaults)
                    exported.append(str(xml_file))
            for sub in vbf.Folders:
                for sc in sub.VBScripts:
                    safe_name = sc.Name.replace(" ", "_").replace("/", "_")
                    if hasattr(sc, "Export"):
                        xml_file = out_dir / f"{sub.Name}_{safe_name}.xml"
                        if xml_file.exists(): xml_file.unlink()
                        sc.Export(FileInfo(str(xml_file)), eng.ExportOptions.WithDefaults)
                        exported.append(str(xml_file))

        if not exported:
            raise TiaError("NO_SCRIPTS",
                f"HMI '{device_name}' hat keine Scripts oder Export nicht verfügbar.", True)
        return {"status": "ok", "device": device_name, "hmi_type": ht,
                "exported": exported, "count": len(exported)}
    return sta.run(_tia_call, _run)

def import_hmi_scripts(device_name, file_path):
    """HMI-Scripts importieren. Unified: Scripts.Import(DirectoryInfo). Advanced HmiTarget: VBScriptComposition.Import(FileInfo, ImportOptions)."""
    def _run():
        _sess.ensure_project()
        import Siemens.Engineering as eng
        from System.IO import FileInfo, DirectoryInfo
        sw, ht = _get_hmi(device_name)
        p = Path(file_path)
        if not p.exists():
            raise TiaError("FILE_NOT_FOUND", f"Datei/Ordner nicht gefunden: {file_path}", True)

        # Unified: Scripts.Import(DirectoryInfo)
        scripts = getattr(sw, "Scripts", None)
        if scripts is not None and hasattr(scripts, "Import"):
            arg = DirectoryInfo(str(p)) if p.is_dir() else FileInfo(str(p))
            scripts.Import(arg)
            return {"status": "ok", "device": device_name, "imported_from": file_path}

        # Advanced HmiTarget: VBScriptComposition.Import(FileInfo, ImportOptions)
        if hasattr(sw, "VBScriptFolder"):
            vbf = sw.VBScriptFolder
            if hasattr(vbf, "VBScripts") and hasattr(vbf.VBScripts, "Import"):
                imported = []
                if p.is_dir():
                    # Alle XML-Dateien im Ordner einzeln importieren
                    for xml_file in p.glob("*.xml"):
                        vbf.VBScripts.Import(FileInfo(str(xml_file)), eng.ImportOptions.Override)
                        imported.append(str(xml_file))
                else:
                    vbf.VBScripts.Import(FileInfo(str(p)), eng.ImportOptions.Override)
                    imported.append(str(p))
                return {"status": "ok", "device": device_name,
                        "imported_from": file_path, "files": imported}

        raise TiaError("SCRIPTS_IMPORT_NOT_SUPPORTED",
            f"Scripts-Import nicht verfügbar für '{device_name}' ({ht}).", False)
    return sta.run(_tia_call, _run)



def import_hmi_screen(device_name, file_path):
    """
    HMI-Screen importieren.
    Advanced: existierenden Screen gleichen Namens löschen, dann Import via API.
    Unified:  ⏭ V21-Limitation — kein Screen-Import via API.
    """
    def _run():
        _sess.ensure_project()
        import Siemens.Engineering as eng
        from System.IO import FileInfo
        sw, ht = _get_hmi(device_name)
        src = Path(file_path)
        if not src.exists():
            raise TiaError("FILE_NOT_FOUND", f"Datei nicht gefunden: {file_path}", True)

        if ht == "Unified":
            raise TiaError("SCREEN_IMPORT_NOT_SUPPORTED",
                f"Screen-Import für Unified nicht verfügbar (V21-Limitation).", False,
                {"hint": "Unified-Screens können nur manuell in TIA Portal importiert werden."})

        # Advanced: Screen-Name aus XML lesen (nicht aus Dateinamen)
        try:
            import xml.etree.ElementTree as _et
            tree = _et.parse(str(src))
            root = tree.getroot()
            # TIA XML: <SW.Screens.Screen ... Name="..."> oder Attribut in ObjectList
            screen_name = None
            for elem in root.iter():
                if "Screen" in elem.tag and elem.get("Name"):
                    screen_name = elem.get("Name")
                    break
                # Alternative: AttributeList/Name
                if elem.tag.endswith("Name") and elem.text:
                    parent_tag = elem.getparent().tag if hasattr(elem, "getparent") else ""
                    if "Screen" in parent_tag:
                        screen_name = elem.text
                        break
            # Fallback: alle Children nach Name-Tag suchen
            if not screen_name:
                for elem in root.iter():
                    if elem.tag == "Name" and elem.text:
                        screen_name = elem.text
                        break
        except Exception:
            screen_name = None

        # Existierenden Screen löschen
        deleted = []
        if screen_name:
            for s in _hmi_screens(sw):
                if s.Name == screen_name:
                    s.Delete()
                    deleted.append(screen_name)
                    break

        _hmi_screens_import(sw, FileInfo(str(src)), eng)
        return {"status": "ok", "device": device_name, "hmi_type": ht,
                "imported_from": file_path, "method": "api_import",
                "deleted_before_import": deleted}
    return sta.run(_tia_call, _run)


def export_hmi_alarms(device_name, output_path=None):
    """
    HMI-Alarme als XML exportieren.
    Gegenstück zu list_hmi_alarms / import_hmi_alarms.
    output_path: Zieldatei (Standard: C:\\tia-mcp\\export\\hmi_alarms_<device>.xml).
    """
    def _run():
        _sess.ensure_project()
        import Siemens.Engineering as eng
        from System.IO import FileInfo
        sw, ht = _get_hmi(device_name)
        out_dir  = _export_dir(None if output_path and Path(output_path).suffix else output_path)
        xml_file = Path(output_path) if (output_path and Path(output_path).suffix) \
                   else out_dir / f"hmi_alarms_{device_name}.xml"
        xml_file.parent.mkdir(parents=True, exist_ok=True)
        if xml_file.exists(): xml_file.unlink()
        # WinCC Advanced: DiscreteAlarms.Export / AnalogAlarms.Export
        # WinCC Unified:  Alarms.Export
        exported = False
        for attr in ("DiscreteAlarms", "AnalogAlarms", "Alarms"):
            col = getattr(sw, attr, None)
            if col and hasattr(col, "Export"):
                col.Export(FileInfo(str(xml_file)), eng.ExportOptions.WithDefaults)
                exported = True
                break
        # Fallback: ganzes HMI-Objekt exportieren
        if not exported and hasattr(sw, "Export"):
            sw.Export(FileInfo(str(xml_file)), eng.ExportOptions.WithDefaults)
            exported = True
        if not exported:
            raise TiaError("ALARM_EXPORT_NOT_SUPPORTED",
                f"Kein Alarm-Export für HMI '{device_name}' ({ht}) verfügbar.", False)
        return {"status": "ok", "device": device_name, "hmi_type": ht,
                "xml_path": str(xml_file)}
    return sta.run(_tia_call, _run)

def import_hmi_alarms(device_name, file_path):
    """
    HMI-Alarme aus XML importieren.
    Gegenstück zu export_hmi_alarms.
    """
    def _run():
        _sess.ensure_project()
        import Siemens.Engineering as eng
        from System.IO import FileInfo
        sw, ht = _get_hmi(device_name)
        fi = FileInfo(file_path)
        if not fi.Exists:
            raise TiaError("FILE_NOT_FOUND", f"Datei nicht gefunden: {file_path}", True)
        imported = False
        for attr in ("DiscreteAlarms", "AnalogAlarms", "Alarms"):
            col = getattr(sw, attr, None)
            if col and hasattr(col, "Import"):
                col.Import(fi, eng.ImportOptions.Override)
                imported = True
                break
        if not imported:
            raise TiaError("ALARM_IMPORT_NOT_SUPPORTED",
                f"Kein Alarm-Import für HMI '{device_name}' ({ht}) verfügbar.", False)
        return {"status": "ok", "device": device_name, "imported_from": file_path}
    return sta.run(_tia_call, _run)

def export_hmi_textlists(device_name, output_path=None):
    """
    HMI-Textlisten als XML exportieren.
    Gegenstück zu list_hmi_textlists / import_hmi_textlists.
    output_path: Zieldatei (Standard: C:\\tia-mcp\\export\\hmi_textlists_<device>.xml).
    """
    def _run():
        _sess.ensure_project()
        import Siemens.Engineering as eng
        from System.IO import FileInfo
        sw, ht = _get_hmi(device_name)
        out_dir  = _export_dir(None if output_path and Path(output_path).suffix else output_path)
        xml_file = Path(output_path) if (output_path and Path(output_path).suffix) \
                   else out_dir / f"hmi_textlists_{device_name}.xml"
        xml_file.parent.mkdir(parents=True, exist_ok=True)
        if xml_file.exists(): xml_file.unlink()
        tl = getattr(sw, "TextLists", None)
        if not tl or not hasattr(tl, "Export"):
            raise TiaError("TEXTLIST_EXPORT_NOT_SUPPORTED",
                f"Kein Textlisten-Export für HMI '{device_name}' ({ht}) verfügbar.", False)
        tl.Export(FileInfo(str(xml_file)), eng.ExportOptions.WithDefaults)
        return {"status": "ok", "device": device_name, "hmi_type": ht,
                "xml_path": str(xml_file)}
    return sta.run(_tia_call, _run)

def import_hmi_textlists(device_name, file_path):
    """
    HMI-Textlisten aus XML importieren.
    Gegenstück zu export_hmi_textlists.
    """
    def _run():
        _sess.ensure_project()
        import Siemens.Engineering as eng
        from System.IO import FileInfo
        sw, ht = _get_hmi(device_name)
        fi = FileInfo(file_path)
        if not fi.Exists:
            raise TiaError("FILE_NOT_FOUND", f"Datei nicht gefunden: {file_path}", True)
        tl = getattr(sw, "TextLists", None)
        if not tl or not hasattr(tl, "Import"):
            raise TiaError("TEXTLIST_IMPORT_NOT_SUPPORTED",
                f"Kein Textlisten-Import für HMI '{device_name}' ({ht}) verfügbar.", False)
        tl.Import(fi, eng.ImportOptions.Override)
        return {"status": "ok", "device": device_name, "imported_from": file_path}
    return sta.run(_tia_call, _run)

def create_hmi_structure(device_name, structure):
    """
    Legt Ordner unter Bilder (Screens) und Tag-Tabellen im HMI an.
    Funktioniert für WinCC Advanced und WinCC Unified.

    structure: dict mit optionalen Keys 'screens' und 'tag_tables'.
    Jeder Wert ist eine Liste aus Strings oder Dicts {"Ordner": ["Kind1","Kind2"]}.

    Beispiel:
        {
          "screens":    ["Start", {"Antriebe": ["Motor1","Motor2"]}, "Pumpen"],
          "tag_tables": ["Allgemein", {"Antriebe": ["Motoren","Umrichter"]}]
        }
    """
    def _flatten(items):
        """Gibt Liste von (parent, child_or_None) zurück."""
        out = []
        for item in items:
            if isinstance(item, str):
                out.append((item, None))
            elif isinstance(item, dict):
                for folder, children in item.items():
                    if not children:
                        out.append((folder, None))
                    else:
                        for child in (children if isinstance(children, list) else [children]):
                            out.append((folder, child))
        return out

    def _ensure(collection, name, child_attr):
        """Gibt Ordner zurück, legt ihn an falls nötig."""
        for f in collection:
            if f.Name == name:
                return f
        return collection.Create(name)

    def _run():
        _sess.ensure_project()
        sw, ht = _get_hmi(device_name)
        created = {"screens": [], "tag_tables": []}
        errors  = []

        # ── Bilder / Screens ────────────────────────────────────────────────
        screen_items = structure.get("screens", [])
        if screen_items:
            folders, ftype = _hmi_screen_folders(sw)
            if folders is None:
                errors.append("Screen-Ordner nicht unterstützt auf diesem HMI-Typ.")
            else:
                # Unterordner-Attribut je nach Typ
                sub_attr = ftype  # "ScreenGroups" oder "ScreenFolders"
                for parent, child in _flatten(screen_items):
                    try:
                        pf = _ensure(folders, parent, sub_attr)
                        entry = f"{parent}"
                        if child:
                            sub = getattr(pf, sub_attr, None)
                            if sub is not None:
                                _ensure(sub, child, sub_attr)
                                entry = f"{parent}/{child}"
                            else:
                                errors.append(f"Unterordner '{parent}/{child}': "
                                              f"keine Sub-Collection ({sub_attr}) gefunden.")
                        if entry not in created["screens"]:
                            created["screens"].append(entry)
                    except Exception as e:
                        errors.append(f"Screen-Ordner '{parent}': {e}")

        # ── Tag-Tabellen ─────────────────────────────────────────────────────
        tag_items = structure.get("tag_tables", [])
        if tag_items:
            tag_root, ttype = _hmi_tag_folders(sw)
            if tag_root is None:
                errors.append("Tag-Tabellen-Ordner nicht unterstützt auf diesem HMI-Typ.")
            else:
                # Advanced V21: TagTableGroups ist die Root-Collection
                # Advanced älter / Unified: TagTableGroup ist ein einzelnes Root-Objekt
                #   dessen Unterordner über .TagTableGroups erreichbar sind
                if ttype == "TagTableGroups":
                    root_coll = tag_root          # direkt iterable Collection
                    sub_attr  = "TagTableGroups"
                else:
                    # TagTableGroup (singular) → Unterordner via .TagTableGroups
                    root_coll = getattr(tag_root, "TagTableGroups", None)
                    sub_attr  = "TagTableGroups"
                    if root_coll is None:
                        errors.append("TagTableGroup hat keine TagTableGroups-Unterordner.")
                        root_coll = []

                for parent, child in _flatten(tag_items):
                    try:
                        pf = _ensure(root_coll, parent, sub_attr)
                        entry = f"{parent}"
                        if child:
                            sub = getattr(pf, sub_attr, None)
                            if sub is not None:
                                _ensure(sub, child, sub_attr)
                                entry = f"{parent}/{child}"
                            else:
                                errors.append(f"Unterordner '{parent}/{child}': "
                                              f"keine Sub-Collection gefunden.")
                        if entry not in created["tag_tables"]:
                            created["tag_tables"].append(entry)
                    except Exception as e:
                        errors.append(f"Tag-Tabellen-Ordner '{parent}': {e}")

        return {
            "status":  "ok" if not errors else "partial",
            "device":  device_name,
            "hmi_type": ht,
            "created": created,
            "errors":  errors,
        }
    return sta.run(_tia_call, _run)

def set_plc_block_source(device_name, block_name, scl_source):
    """
    SCL-Quellcode direkt als String in einen Baustein schreiben und importieren.
    Gegenstück zu get_plc_block_source.
    scl_source: SCL-Code als String — wird als temporäre XML-Datei verpackt und importiert.
    Voraussetzung: Baustein muss bereits existieren (gleicher Name, gleicher Typ).
    """
    def _run():
        _sess.ensure_project()
        import Siemens.Engineering as eng
        from System.IO import FileInfo
        plc = _find_sw(_sess.project, device_name, "PlcSoftware", eng)
        if not plc:
            raise TiaError("PLC_NOT_FOUND", f"PLC '{device_name}' nicht gefunden.", True)
        block = _find_block(plc, block_name)
        if not block:
            avail = [b.Name for b in plc.BlockGroup.Blocks]
            raise TiaError("BLOCK_NOT_FOUND", f"Baustein '{block_name}' nicht gefunden.", True,
                           {"available": avail})
        lang = str(getattr(block, "ProgrammingLanguage", ""))
        if "SCL" not in lang.upper() and "StructuredControlLanguage" not in lang:
            raise TiaError("WRONG_LANGUAGE",
                f"Baustein '{block_name}' ist {lang}, kein SCL — set_plc_block_source nur für SCL.", True)

        # Aktuelles XML exportieren und SCL-Body ersetzen
        out_dir  = _export_dir()
        xml_file = out_dir / f"{block_name}_set.xml"
        if xml_file.exists(): xml_file.unlink()
        block.Export(FileInfo(str(xml_file)), eng.ExportOptions.WithDefaults)
        xml_content = xml_file.read_text(encoding="utf-8")

        import re as _re

        # Höchste vorhandene UId im XML ermitteln
        existing_uids = [int(u) for u in _re.findall(r'UId="(\d+)"', xml_content)]
        uid_start = max(existing_uids) + 1 if existing_uids else 100

        def _scl_to_tokens(code: str, uid: int) -> str:
            """SCL-Code zeilenweise in Token-Elemente mit UId umwandeln."""
            def esc(s):
                return (s.replace("&", "&amp;")
                          .replace("<", "&lt;")
                          .replace(">", "&gt;")
                          .replace('"', "&quot;")
                          .replace("\r\n", "&#xD;&#xA;")
                          .replace("\n",   "&#xD;&#xA;")
                          .replace("\r",   "&#xD;&#xA;"))
            parts = []
            for line in code.splitlines(keepends=True):
                e = esc(line)
                if e:
                    parts.append(f'<Token Text="{e}" UId="{uid}" />')
                    uid += 1
            return "\n              ".join(parts)

        new_xml  = xml_content
        replaced = False
        token_xml = _scl_to_tokens(scl_source, uid_start)

        # V21 Self-Closing: <StructuredText ... />
        sc_pat = r'<StructuredText(\b[^>]*)\/>'
        m_sc   = _re.search(sc_pat, new_xml)
        if m_sc:
            attrs = m_sc.group(1)
            replacement = f'<StructuredText{attrs}>\n              {token_xml}\n            </StructuredText>'
            new_xml = _re.sub(sc_pat, replacement, new_xml, count=1)
            replaced = True

        # V21 Container mit vorhandenen Tokens ersetzen
        if not replaced:
            cont_pat = r'(<StructuredText\b[^>]*>)(.*?)(</StructuredText>)'
            if _re.search(cont_pat, new_xml, _re.DOTALL):
                new_xml = _re.sub(cont_pat,
                    lambda m: m.group(1) + f'\n              {token_xml}\n            ' + m.group(3),
                    new_xml, flags=_re.DOTALL, count=1)
                replaced = True

        # V19/V20 Fallback
        if not replaced:
            for tag in ["Body", "SourceText"]:
                pat = rf"(<{tag}[^>]*>)(.*?)(</{tag}>)"
                if _re.search(pat, new_xml, _re.DOTALL):
                    new_xml = _re.sub(pat,
                        lambda x: x.group(1) + scl_source + x.group(3),
                        new_xml, flags=_re.DOTALL)
                    replaced = True
                    break

        if not replaced:
            raise TiaError("SCL_TAG_NOT_FOUND",
                f"Kein SCL-Tag in XML von '{block_name}' gefunden.", False)

        xml_file.write_text(new_xml, encoding="utf-8")

        # Import — Baustein liegt ggf. in Untergruppe
        def _find_group(plc_sw, bname):
            stack = [plc_sw.BlockGroup]
            while stack:
                grp = stack.pop()
                for b in grp.Blocks:
                    if b.Name == bname:
                        return grp
                for sub in grp.Groups:
                    stack.append(sub)
            return plc_sw.BlockGroup

        target_group = _find_group(plc, block_name)
        result = target_group.Blocks.Import(FileInfo(str(xml_file)), eng.ImportOptions.Override)
        _log("plc").info(f"SCL gesetzt: {block_name}")
        return {"status": "ok", "block": block_name, "language": lang,
                "xml_path": str(xml_file),
                "blocks": [str(b) for b in result] if result else []}
    return sta.run(_tia_call, _run)

# ═══════════════════════════════════════════════════════════════════════════════
# DATEI-HILFSFUNKTIONEN
# ═══════════════════════════════════════════════════════════════════════════════

def write_import_file(filename: str, content: str) -> dict:
    """
    Schreibt Dateiinhalt (z.B. aus Chat-Upload) in den Import-Ordner.
    Claude kann so hochgeladene XML-Dateien fuer den Import bereitstellen.
    """
    import_dir = Path(_DEFAULT_EXPORT) / "import"
    import_dir.mkdir(parents=True, exist_ok=True)
    out = import_dir / filename
    out.write_text(content, encoding="utf-8")
    _log("import").info(f"Import-Datei geschrieben: {out}")
    return {"status": "ok", "path": str(out)}

def read_export_file(file_path: str) -> dict:
    """Liest eine exportierte Datei und gibt den Inhalt zurueck."""
    p = Path(file_path)
    if not p.exists():
        raise TiaError("FILE_NOT_FOUND", f"Datei nicht gefunden: {file_path}", True)
    content = p.read_text(encoding="utf-8")
    return {"status": "ok", "path": file_path,
            "content": content, "size_kb": round(len(content)/1024, 1)}

# ═══════════════════════════════════════════════════════════════════════════════
# KOMPILIEREN
# ═══════════════════════════════════════════════════════════════════════════════

def compile_plc(device_name):
    """SPS kompilieren — behebt inkonsistente Bausteine vor dem Export."""
    def _run():
        _sess.ensure_project()
        import Siemens.Engineering as eng
        plc = _find_sw(_sess.project, device_name, "PlcSoftware", eng)
        if not plc:
            raise TiaError("PLC_NOT_FOUND", f"PLC '{device_name}' nicht gefunden.", True)

        # ICompilable liegt in V21 in Siemens.Engineering.Base (Compiler-Namespace).
        # Strategie: alle geladenen Assemblies nach dem Typ durchsuchen.
        compilable_type = None
        step7_asm = plc.GetType().Assembly
        # 1) Direkt aus Step7-Assembly
        compilable_type = step7_asm.GetType("Siemens.Engineering.Compiler.ICompilable")
        # 2) Aus Base-Assembly (referenziert von Step7)
        if not compilable_type:
            for ref in step7_asm.GetReferencedAssemblies():
                if "Base" in str(ref.Name):
                    try:
                        import System.Reflection as refl
                        base_asm = refl.Assembly.Load(ref)
                        compilable_type = base_asm.GetType(
                            "Siemens.Engineering.Compiler.ICompilable")
                        if compilable_type:
                            break
                    except Exception:
                        pass
        # 3) Fallback: Namespace direkt importieren
        if not compilable_type:
            try:
                from Siemens.Engineering.Compiler import ICompilable
                compilable_type = ICompilable
            except Exception:
                pass

        if not compilable_type:
            raise TiaError("COMPILE_NOT_SUPPORTED",
                "ICompilable nicht gefunden — Siemens.Engineering.Base.dll pruefen.", False)

        compiler = plc.GetService[compilable_type]()
        if not compiler:
            raise TiaError("COMPILE_NOT_SUPPORTED",
                "Kein Compiler-Service verfuegbar.", False)

        result = compiler.Compile()
        status = "ok" if result.ErrorCount == 0 else "error"
        _log("compile").info(
            f"Kompiliert {device_name}: {result.ErrorCount} Fehler, {result.WarningCount} Warnungen")

        messages = []
        try:
            for msg in result.Messages:
                messages.append({
                    "severity":    str(msg.Severity),
                    "description": str(msg.Description),
                    "path":        str(getattr(msg, "Path", ""))
                })
        except Exception:
            pass

        return {
            "status":   status,
            "device":   device_name,
            "errors":   result.ErrorCount,
            "warnings": result.WarningCount,
            "messages": messages[:20]
        }
    return sta.run(_tia_call, _run, timeout=_STA_TIMEOUT_HEAVY)