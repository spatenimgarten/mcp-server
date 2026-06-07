"""
tia_export.py — DB-Variablen mit Querverweis-Analyse
======================================================
Liest alle globalen DBs aus TIA Portal, exportiert alle Bausteine als XML,
analysiert Zugriffe (Read / Write / ReadWrite) und schreibt eine Excel-Datei
zur manuellen Prüfung vor dem HMI-Import.

Voraussetzungen:
  TIA Portal V21 offen, Projekt geladen
  pip install openpyxl

Aufruf:
  .venv\\Scripts\\python.exe tools\\tia_export.py
  .venv\\Scripts\\python.exe tools\\tia_export.py --device PLC_1 --out C:\\tia-mcp\\export\\projekt.xlsx

Ergebnis:
  projekt.xlsx mit Sheet "DB_Variablen":
    DB | Gruppe | Variable | Datentyp | Zugriff | Ins HMI? | HMI-Tagname | Tabelle | Hi | Lo | Einheit | Archiv | Zykluszeit_ms
"""

import sys, os, re, argparse
from pathlib import Path
from collections import defaultdict

# ── Pfad zum MCP-Server-Modul ──────────────────────────────────────────────────
_SCRIPT_DIR = Path(__file__).resolve().parent
_ROOT = _SCRIPT_DIR.parent          # tools/ liegt unter dem Repo-Root
if str(_ROOT) not in sys.path:
    sys.path.insert(0, str(_ROOT))

import tia
from tia import TiaError

# ── Konfiguration ──────────────────────────────────────────────────────────────

DEFAULT_DEVICE  = ""                           # leer = erstes PLC im Projekt
DEFAULT_OUT     = r"C:\tia-mcp\export\projekt.xlsx"
EXPORT_TMP      = r"C:\tia-mcp\export\xml_tmp"   # temporärer XML-Exportordner

# Datentypen die als "analog" / HMI-relevant gelten → werden als Read vormarkiert
ANALOG_TYPES = {"Real", "LReal", "Int", "DInt", "UInt", "UDInt",
                "SInt", "USInt", "Word", "DWord", "LWord", "Byte"}

# ── Logging ────────────────────────────────────────────────────────────────────

def log(msg):
    print(f"  {msg}", flush=True)


# ══════════════════════════════════════════════════════════════════════════════
# SCHRITT 1 — DB-Variablen lesen
# ══════════════════════════════════════════════════════════════════════════════

def read_db_variables(device_name: str) -> list[dict]:
    """
    Liest alle globalen DBs und deren Interface-Variablen via execute_openness.
    Gibt eine flache Liste zurück:
      {"db": str, "variable": str, "datatype": str}
    """
    code = """
stack = [(plc.BlockGroup, "")]
result = []
while stack:
    group, path = stack.pop()
    for block in group.Blocks:
        btype = type(block).__name__
        # Nur DataBlocks — keine Instanz-DBs (InstanceDataBlock)
        if btype != "DataBlock":
            continue
        db_name = block.Name
        try:
            members = block.Interface.Members
            for m in members:
                mname = safe_str(m.Name)
                mtype = safe_str(getattr(m, "Datatype", ""))
                # UDT-Member auffalten (eine Ebene)
                try:
                    sub_members = list(m.Members)
                    if sub_members:
                        for sm in sub_members:
                            smname = safe_str(sm.Name)
                            smtype = safe_str(getattr(sm, "Datatype", ""))
                            result.append({
                                "db": db_name,
                                "gruppe": path,
                                "variable": f"{mname}.{smname}",
                                "datatype": smtype,
                            })
                        continue
                except Exception:
                    pass
                result.append({
                    "db": db_name,
                    "gruppe": path,
                    "variable": mname,
                    "datatype": mtype,
                })
        except Exception as e:
            result.append({"db": db_name, "gruppe": path,
                           "variable": f"[FEHLER: {e}]", "datatype": ""})
    for sub in group.Groups:
        sub_path = f"{path}/{sub.Name}".lstrip("/")
        stack.append((sub, sub_path))
"""

    # find_software mit device_name oder erstem PLC
    setup = f"""
plc = find_software("{device_name}", "PlcSoftware")
if not plc:
    plc = find_software("", "PlcSoftware")
if not plc:
    result = {{"error": "PLC nicht gefunden"}}
else:
"""
    # Einrücken des Haupt-Codes
    indented = "\n".join("    " + line for line in code.strip().splitlines())
    full_code = setup + indented

    res = tia.execute_openness(full_code, mode="read")
    if isinstance(res.get("result"), dict) and "error" in res["result"]:
        raise RuntimeError(res["result"]["error"])
    return res.get("result") or []


# ══════════════════════════════════════════════════════════════════════════════
# SCHRITT 2 — Alle Bausteine exportieren
# ══════════════════════════════════════════════════════════════════════════════

def export_all_blocks(device_name: str, export_dir: str) -> list[str]:
    """
    Exportiert alle OBs, FCs, FBs (nicht DBs) als XML in export_dir.
    Gibt Liste der erzeugten XML-Dateien zurück.
    """
    Path(export_dir).mkdir(parents=True, exist_ok=True)

    code = f"""
from System.IO import FileInfo, DirectoryInfo
import Siemens.Engineering as _eng

plc = find_software("{device_name}", "PlcSoftware")
if not plc:
    plc = find_software("", "PlcSoftware")

exported = []
errors   = []

stack = [plc.BlockGroup]
while stack:
    group = stack.pop()
    for block in group.Blocks:
        btype = type(block).__name__
        # DBs überspringen — nur Programm-Bausteine
        if "DataBlock" in btype:
            continue
        bname = safe_str(block.Name)
        xml_path = r"{export_dir}\\" + bname + ".xml"
        try:
            block.Export(FileInfo(xml_path), _eng.ExportOptions.WithDefaults)
            exported.append(xml_path)
        except Exception as e:
            errors.append({{"block": bname, "error": str(e)}})
    for sub in group.Groups:
        stack.append(sub)

result = {{"exported": exported, "errors": errors}}
"""

    res = tia.execute_openness(code, mode="read")
    data = res.get("result", {})
    exported = data.get("exported", [])
    errors   = data.get("errors", [])

    if errors:
        for e in errors:
            log(f"  ⚠ Export-Fehler {e['block']}: {e['error']}")
    log(f"  {len(exported)} Bausteine exportiert nach {export_dir}")
    return exported


# ══════════════════════════════════════════════════════════════════════════════
# SCHRITT 3 — XML-Dateien analysieren → Zugriffsliste
# ══════════════════════════════════════════════════════════════════════════════

# Zugriffs-Typen
READ      = "Read"
WRITE     = "Write"
READWRITE = "ReadWrite"

def _merge_access(existing: str | None, new: str) -> str:
    if existing is None:
        return new
    if existing == new:
        return existing
    return READWRITE


def _parse_xml_file(xml_path: str) -> dict[tuple[str, str], str]:
    """
    Parst eine exportierte Baustein-XML und gibt zurück:
      {(db_name, variable_name): "Read" | "Write" | "ReadWrite"}

    Abgedeckte Fälle:
      - SCL: direkte Memberzugriffe  "DB_Name".Variable
      - FUP/LAD: <Access Scope="GlobalVariable"> Knoten
      - Slice-Zugriffe (%X, %B, %W)
      - Konstante Array-Indizes
    """
    try:
        content = Path(xml_path).read_text(encoding="utf-8", errors="replace")
    except Exception:
        return {}

    accesses: dict[tuple[str, str], str] = {}

    # ── FUP/LAD XML-Analyse ───────────────────────────────────────────────────
    # Muster: <Access Scope="GlobalVariable" ...>
    #           <Symbol>
    #             <Component Name="DB_Motor"/>
    #             <Component Name="Drehzahl"/>
    #           </Symbol>
    #         </Access>
    # Das Attribut "Informational" kennzeichnet reine Kommentar-Zugriffe → ignorieren
    if "<Access " in content:
        # Alle Access-Blöcke extrahieren
        access_blocks = re.findall(
            r'<Access\s[^>]*Scope="GlobalVariable"[^>]*>(.*?)</Access>',
            content, re.DOTALL
        )
        for block in access_blocks:
            # Informational-Zugriffe (nur Anzeigehinweise) ignorieren
            if 'Informational="true"' in block:
                continue

            # Komponenten extrahieren
            components = re.findall(r'<Component\s+Name="([^"]+)"', block)
            if len(components) < 2:
                continue

            db_name  = components[0]
            var_name = components[1]

            # Array-Index anfügen wenn vorhanden (konstanter Index)
            index_m = re.search(r'<Component\s+Name="\[(\d+)\]"', block)
            if index_m:
                var_name = f"{var_name}[{index_m.group(1)}]"

            # Slice-Zugriff (%X3, %B0 etc.)
            slice_m = re.search(r'<Access\s[^>]*Scope="LocalVariable"[^>]*>\s*<Symbol>\s*<Component Name="%([XBW])(\d+)"/>', block)
            # (Slice nur als Hinweis — kein eigener Key)

            # Write-Erkennung: Block liegt in einer Zuweisung / Spule
            # Heuristik: übergeordnetes Element prüfen
            # Im XML-Kontext: wenn Access in einem <Assignment> oder <Coil> steht → Write
            # Einfachere Heuristik: Wenn der Block von einem Write-Indikator umgeben ist
            is_write = _is_write_context_xml(content, block)

            access_type = WRITE if is_write else READ
            key = (db_name, var_name)
            accesses[key] = _merge_access(accesses.get(key), access_type)

    # ── SCL-Analyse ───────────────────────────────────────────────────────────
    # SCL wird in <StructuredText> oder ähnlichen Tags gespeichert
    scl_matches = re.findall(
        r'<(?:StructuredText|Body|SourceText)>(.*?)</(?:StructuredText|Body|SourceText)>',
        content, re.DOTALL
    )
    for scl_block in scl_matches:
        # CDATA entfernen
        scl = re.sub(r'<!\[CDATA\[(.*?)\]\]>', r'\1', scl_block, flags=re.DOTALL)
        _parse_scl(scl, accesses)

    return accesses


def _is_write_context_xml(full_xml: str, access_block: str) -> bool:
    """
    Heuristik: Prüft ob ein Access-Block in einem schreibenden Kontext steht.
    Write-Kontexte in TIA FUP/LAD XML:
      - <Assignment> ... </Assignment>  (Zuweisung)
      - <Coil> ... </Coil>              (Ausgangs-Spule)
      - <SetCoil>, <ResetCoil>
      - <Call> mit Output-Parameter     (schwieriger)
    """
    # Position des Blocks im Gesamt-XML finden
    pos = full_xml.find(access_block[:60])  # ersten 60 Zeichen als Anker
    if pos < 0:
        return False

    # 500 Zeichen vor dem Access-Block nach Write-Indikatoren suchen
    context_before = full_xml[max(0, pos - 500):pos]
    write_indicators = ["<Assignment", "<Coil", "<SetCoil", "<ResetCoil",
                        "<MoveInstruction", 'Direction="Output"']
    for indicator in write_indicators:
        if indicator in context_before:
            # Prüfen ob der Indikator noch "offen" ist (kein schließendes Tag)
            tag_name = indicator.lstrip("<").split()[0].rstrip(">")
            close_tag = f"</{tag_name}>"
            last_open  = context_before.rfind(indicator)
            last_close = context_before.rfind(close_tag)
            if last_open > last_close:
                return True
    return False


def _parse_scl(scl: str, accesses: dict) -> None:
    """
    Analysiert SCL-Quellcode auf DB-Zugriffe.

    Write:    "DB".Var :=  oder  "DB".Var[i] :=
    Read:     alles andere wo "DB".Var vorkommt
    """
    # Muster: "DB_Name".Member  (optional: .SubMember oder [Index])
    # Gruppen: 1=DB-Name, 2=Variable (inkl. Subpath und Index)
    pattern = re.compile(
        r'"([A-Za-z0-9_]+)"\s*\.\s*([A-Za-z0-9_]+(?:\.[A-Za-z0-9_]+)*(?:\[\d+\])?)',
        re.IGNORECASE
    )

    for m in pattern.finditer(scl):
        db_name  = m.group(1)
        var_name = m.group(2)

        # Write-Erkennung: nach dem Match folgt ":=" (mit optionalem Whitespace)
        after = scl[m.end():m.end() + 10].lstrip()
        is_write = after.startswith(":=")

        access_type = WRITE if is_write else READ
        key = (db_name, var_name)
        accesses[key] = _merge_access(accesses.get(key), access_type)


def analyze_accesses(xml_files: list[str]) -> dict[tuple[str, str], str]:
    """Analysiert alle XML-Dateien und aggregiert Zugriffe."""
    all_accesses: dict[tuple[str, str], str] = {}
    for xml_file in xml_files:
        file_accesses = _parse_xml_file(xml_file)
        for key, access in file_accesses.items():
            all_accesses[key] = _merge_access(all_accesses.get(key), access)
    return all_accesses


# ══════════════════════════════════════════════════════════════════════════════
# SCHRITT 4 — Excel schreiben
# ══════════════════════════════════════════════════════════════════════════════

def write_excel(
    db_variables: list[dict],
    accesses: dict[tuple[str, str], str],
    output_path: str
) -> None:
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise RuntimeError("openpyxl nicht installiert. Bitte: pip install openpyxl")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "DB_Variablen"

    # ── Spalten ───────────────────────────────────────────────────────────────
    COLS = [
        ("DB",            18),
        ("Gruppe",        22),
        ("Variable",      28),
        ("Datentyp",      14),
        ("Zugriff",       14),
        ("Ins HMI?",      10),
        ("HMI-Tagname",   28),
        ("Tabelle",       16),
        ("Hi",            10),
        ("Lo",            10),
        ("Einheit",       10),
        ("Archiv",        10),
        ("Zykluszeit_ms", 14),
    ]

    # ── Farben ────────────────────────────────────────────────────────────────
    COLOR_HEADER    = "1F4E79"   # Dunkelblau
    COLOR_READ      = "E2EFDA"   # Hellgrün  → ins HMI vormarkiert
    COLOR_WRITE     = "FCE4D6"   # Lachs     → nicht ins HMI
    COLOR_READWRITE = "FFF2CC"   # Gelb      → manuell prüfen
    COLOR_NONE      = "F2F2F2"   # Hellgrau  → nicht verwendet

    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ── Header ────────────────────────────────────────────────────────────────
    header_font  = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
    header_fill  = PatternFill("solid", fgColor=COLOR_HEADER)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.row_dimensions[1].height = 30
    for col_idx, (col_name, col_width) in enumerate(COLS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font   = header_font
        cell.fill   = header_fill
        cell.alignment = header_align
        cell.border = border
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    ws.freeze_panes = "A2"

    # ── Daten ─────────────────────────────────────────────────────────────────
    default_align = Alignment(vertical="center")

    for row_idx, var in enumerate(db_variables, start=2):
        db_name  = var.get("db", "")
        gruppe   = var.get("gruppe", "")
        var_name = var.get("variable", "")
        dtype    = var.get("datatype", "")

        # Zugriff ermitteln — normalisierter DB-Name-Abgleich
        access = accesses.get((db_name, var_name))
        # Fallback: nur ersten Teil bei Sub-Membern versuchen
        if access is None and "." in var_name:
            top = var_name.split(".")[0]
            access = accesses.get((db_name, top))

        access_str = access if access else "—"

        # Vormarkierung "Ins HMI?"
        ins_hmi = "ja" if access_str == READ else ""

        # Zeilenfarbe
        if access_str == READ:
            row_color = COLOR_READ
        elif access_str == WRITE:
            row_color = COLOR_WRITE
        elif access_str == READWRITE:
            row_color = COLOR_READWRITE
        else:
            row_color = COLOR_NONE

        row_fill = PatternFill("solid", fgColor=row_color)

        values = [db_name, gruppe, var_name, dtype, access_str, ins_hmi,
                  "", "", "", "", "", "", ""]

        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.fill      = row_fill
            cell.border    = border
            cell.alignment = default_align
            cell.font      = Font(name="Calibri", size=10)

    # ── AutoFilter ────────────────────────────────────────────────────────────
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLS))}{len(db_variables) + 1}"

    # ── Legende ───────────────────────────────────────────────────────────────
    ws_leg = wb.create_sheet("Legende")
    legend = [
        ("Farbe",     "Zugriff",   "Bedeutung"),
        ("Grün",      "Read",      "Variable wird nur gelesen → ins HMI vormarkiert"),
        ("Lachs",     "Write",     "Variable wird nur geschrieben → kein HMI-Tag nötig"),
        ("Gelb",      "ReadWrite", "Variable wird gelesen UND geschrieben → manuell prüfen"),
        ("Grau",      "—",         "Variable kommt in keinem Baustein vor → evtl. ungenutzt"),
    ]
    fills = [
        PatternFill("solid", fgColor=COLOR_HEADER),
        PatternFill("solid", fgColor=COLOR_READ),
        PatternFill("solid", fgColor=COLOR_WRITE),
        PatternFill("solid", fgColor=COLOR_READWRITE),
        PatternFill("solid", fgColor=COLOR_NONE),
    ]
    for r, (row_data, fill) in enumerate(zip(legend, fills), start=1):
        for c, val in enumerate(row_data, start=1):
            cell = ws_leg.cell(row=r, column=c, value=val)
            cell.fill = fill
            if r == 1:
                cell.font = Font(bold=True, color="FFFFFF", name="Calibri")
            else:
                cell.font = Font(name="Calibri", size=10)
            cell.border = border
    ws_leg.column_dimensions["A"].width = 12
    ws_leg.column_dimensions["B"].width = 14
    ws_leg.column_dimensions["C"].width = 55

    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    log(f"Excel gespeichert: {output_path}")
    log(f"  {len(db_variables)} Variablen, "
        f"{sum(1 for v in db_variables if accesses.get((v['db'], v['variable'])) == READ)} als Read vormarkiert")


# ══════════════════════════════════════════════════════════════════════════════
# MAIN
# ══════════════════════════════════════════════════════════════════════════════

def main():
    parser = argparse.ArgumentParser(description="TIA Export — DB-Variablen mit Querverweisanalyse")
    parser.add_argument("--device", default=DEFAULT_DEVICE,
                        help="DeviceItem-Name der SPS (leer = erstes PLC im Projekt)")
    parser.add_argument("--out", default=DEFAULT_OUT,
                        help="Ausgabepfad für die Excel-Datei")
    parser.add_argument("--xml-dir", default=EXPORT_TMP,
                        help="Temporärer Ordner für XML-Exporte")
    parser.add_argument("--keep-xml", action="store_true",
                        help="XML-Dateien nach Analyse nicht löschen")
    args = parser.parse_args()

    print("\n═══════════════════════════════════════════════════")
    print("  TIA Export — DB-Variablen + Querverweis")
    print("═══════════════════════════════════════════════════\n")

    # ── TIA verbinden ─────────────────────────────────────────────────────────
    log("Verbinde mit TIA Portal...")
    tia.setup()
    try:
        r = tia.connect_portal(mode="attach")
        log(f"Portal: {r.get('tia_version', '?')} (PID {r.get('process_id', '?')})")
    except TiaError as e:
        print(f"\nFEHLER: {e.message}")
        sys.exit(1)

    try:
        r = tia.attach_project()
        log(f"Projekt: {r.get('project', '?')}")
    except TiaError as e:
        print(f"\nFEHLER: {e.message}")
        sys.exit(1)

    # ── Schritt 1: DB-Variablen lesen ─────────────────────────────────────────
    log("\n[1/3] Lese DB-Variablen...")
    try:
        db_vars = read_db_variables(args.device)
        log(f"  {len(db_vars)} Variablen aus {len({v['db'] for v in db_vars})} DBs gelesen")
    except Exception as e:
        print(f"\nFEHLER beim Lesen der DBs: {e}")
        tia.teardown()
        sys.exit(1)

    # ── Schritt 2: Bausteine exportieren ──────────────────────────────────────
    log("\n[2/3] Exportiere Bausteine für Querverweis-Analyse...")
    try:
        xml_files = export_all_blocks(args.device, args.xml_dir)
    except Exception as e:
        print(f"\nFEHLER beim Export: {e}")
        tia.teardown()
        sys.exit(1)

    # ── Schritt 3: Zugriffe analysieren ───────────────────────────────────────
    log("\n[2/3] Analysiere Zugriffe...")
    accesses = analyze_accesses(xml_files)
    log(f"  {len(accesses)} DB-Zugriffe gefunden")

    # Statistik
    read_count  = sum(1 for v in accesses.values() if v == READ)
    write_count = sum(1 for v in accesses.values() if v == WRITE)
    rw_count    = sum(1 for v in accesses.values() if v == READWRITE)
    log(f"  Read: {read_count}  Write: {write_count}  ReadWrite: {rw_count}")

    # Temporäre XML-Dateien aufräumen
    if not args.keep_xml:
        cleaned = 0
        for f in xml_files:
            try:
                Path(f).unlink()
                cleaned += 1
            except Exception:
                pass
        if cleaned:
            log(f"  {cleaned} temporäre XML-Dateien gelöscht (--keep-xml zum Behalten)")

    # ── Schritt 4: Excel schreiben ────────────────────────────────────────────
    log("\n[3/3] Schreibe Excel...")
    try:
        write_excel(db_vars, accesses, args.out)
    except Exception as e:
        print(f"\nFEHLER beim Excel-Export: {e}")
        tia.teardown()
        sys.exit(1)

    tia.teardown()
    print("\n✓ Fertig.\n")


if __name__ == "__main__":
    main()